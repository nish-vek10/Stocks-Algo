# Path: research/experiments/09E_batched_report.py
"""
ALGO-STOCKS Phase 09E — Time-Batched Analysis & Investor Report

PURPOSE
-------
Splits the 09B universe trades into configurable time windows, computes
performance metrics per batch, merges into a combined summary, and writes
a clean investor-ready Excel workbook with per-ticker and per-batch sheets.

BATCH METHODOLOGY — ENTRY-DATE ATTRIBUTION
-------------------------------------------
Each trade is assigned to the batch window in which it ENTERED.
A trade entered in Batch 1 that exits in Batch 2 is counted in Batch 1.
Trades are NEVER force-closed. This is the standard fund reporting approach:
you measure the vintage of each entry decision, not the calendar window.

This means:
  - No artificial cuts or incomplete trade distortions
  - Each batch is a clean snapshot of "decisions made in this period"
  - The combined report = Batch 1 + Batch 2 (no overlaps, no gaps)

TIME WINDOWS (configurable in config/backtest.yaml under 'batches:')
  Batch 1: entry_date in [2022-01-01, 2023-12-31]
  Batch 2: entry_date in [2024-01-01, 2026-12-31]

TICKER FILTERING
----------------
If batches.use_filtered_tickers: true, only tickers that passed the 09D
filter are included. Requires 09D to have been run first.

OUTPUTS
-------
output/reports/batched/
    batch_1_trades.parquet        — all trades in Batch 1
    batch_2_trades.parquet        — all trades in Batch 2
    batch_1_ticker_summary.csv    — per-ticker metrics for Batch 1
    batch_2_ticker_summary.csv    — per-ticker metrics for Batch 2
    combined_ticker_summary.csv   — per-ticker metrics across all batches
    investor_report.xlsx          — multi-sheet investor presentation

RUN FROM PROJECT ROOT:
  python research/experiments/09E_batched_report.py
"""
from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path
import sys
from typing import Optional

import numpy as np
import pandas as pd
import yaml

# ── Project root resolution ───────────────────────────────────────────────────
ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT))

# ══════════════════════════════════════════════════════════════════════════════
# CONFIGURABLE PATHS
# ══════════════════════════════════════════════════════════════════════════════
BACKTEST_CFG    = ROOT / "config" / "backtest.yaml"
BACKTESTS_DIR   = ROOT / "output" / "backtests"
FILTER_DIR      = ROOT / "output" / "reports" / "universe_filter"
REPORTS_DIR     = ROOT / "output" / "reports" / "batched"
MEMBERSHIPS     = ROOT / "data" / "metadata" / "spiders" / "spider_memberships.csv"
# ══════════════════════════════════════════════════════════════════════════════


def _load_cfg(path: Path) -> dict:
    with path.open("r", encoding="utf-8") as f:
        return yaml.safe_load(f) or {}


def _sector_from_spider_id(spider_id) -> str:
    if not isinstance(spider_id, str):
        return "Unknown"
    return spider_id.replace("SECTOR_", "").replace("_", " ").title()


def compute_batch_metrics(trades: pd.DataFrame, batch_name: str) -> dict:
    """
    Compute full performance metrics for a set of trades.
    Uses entry-date attribution — all trades belong to the batch they entered.
    """
    if trades.empty:
        return {"batch_name": batch_name, "error": "no_trades"}

    t          = trades
    total      = len(t)
    wins       = t[t["pnl_dollar"] > 0]
    losses     = t[t["pnl_dollar"] <= 0]
    win_count  = len(wins)
    loss_count = len(losses)

    gross_wins   = float(wins["pnl_dollar"].sum())   if win_count  > 0 else 0.0
    gross_losses = float(losses["pnl_dollar"].abs().sum()) if loss_count > 0 else 0.0
    net_pnl      = float(t["pnl_dollar"].sum())
    pf           = round(gross_wins / gross_losses, 4) if gross_losses > 0 else float("inf")
    exp_r        = float(t["pnl_r"].mean())           if total > 0 else 0.0

    hold = t["hold_days"].dropna()

    # Date range of entries in this batch
    entry_dates = pd.to_datetime(t["entry_date"])
    exit_dates  = pd.to_datetime(t["exit_date"])

    return {
        "batch_name":        batch_name,
        "total_trades":      total,
        "winning_trades":    win_count,
        "losing_trades":     loss_count,
        "win_rate_pct":      round(win_count / total * 100, 2) if total > 0 else 0.0,
        "avg_win_pct":       round(float(t.loc[t["pnl_dollar"] > 0,  "pnl_pct"].mean()), 4) if win_count  > 0 else 0.0,
        "avg_loss_pct":      round(float(t.loc[t["pnl_dollar"] <= 0, "pnl_pct"].mean()), 4) if loss_count > 0 else 0.0,
        "best_trade_pct":    round(float(t["pnl_pct"].max()), 4) if total > 0 else 0.0,
        "worst_trade_pct":   round(float(t["pnl_pct"].min()), 4) if total > 0 else 0.0,
        "expectancy_r":      round(exp_r, 4),
        "profit_factor":     pf,
        "gross_wins_usd":    round(gross_wins, 2),
        "gross_losses_usd":  round(gross_losses, 2),
        "net_pnl_usd":       round(net_pnl, 2),
        "avg_hold_days":     round(float(hold.mean()), 1)  if not hold.empty else 0.0,
        "max_hold_days":     int(hold.max())               if not hold.empty else 0,
        "stage6_entries":    int((t["signal_type"] == "stage6_entry").sum()),
        "stage7_entries":    int((t["signal_type"] == "stage7_entry").sum()),
        "unique_tickers":    int(t["ticker"].nunique()),
        "entry_date_first":  str(entry_dates.min().date()) if not entry_dates.empty else "",
        "entry_date_last":   str(entry_dates.max().date()) if not entry_dates.empty else "",
        "exit_date_last":    str(exit_dates.max().date())  if not exit_dates.empty else "",
        "exit_reasons":      t["exit_reason"].value_counts().to_dict(),
    }


def compute_ticker_metrics_in_batch(
    trades: pd.DataFrame, batch_name: str, batch_label: str
) -> pd.DataFrame:
    """
    Compute per-ticker metrics for trades in a single batch window.
    Returns a DataFrame with one row per ticker.
    """
    if trades.empty:
        return pd.DataFrame()

    rows = []
    for ticker, grp in trades.groupby("ticker"):
        g = grp.copy()
        total = len(g)
        wins  = g[g["pnl_dollar"] > 0]
        loss  = g[g["pnl_dollar"] <= 0]

        gross_w = float(wins["pnl_dollar"].sum())       if len(wins) > 0 else 0.0
        gross_l = float(loss["pnl_dollar"].abs().sum()) if len(loss) > 0 else 0.0
        net_pnl = float(g["pnl_dollar"].sum())
        pf      = round(gross_w / gross_l, 4)           if gross_l > 0 else float("inf")
        exp_r   = float(g["pnl_r"].mean())              if total > 0 else 0.0

        rows.append({
            "ticker":          ticker,
            "sector":          g["sector"].iloc[0]     if "sector" in g.columns else "Unknown",
            "spider_id":       g["spider_id"].iloc[0]  if "spider_id" in g.columns else None,
            "batch":           batch_label,
            "batch_name":      batch_name,
            "total_trades":    total,
            "winning_trades":  len(wins),
            "losing_trades":   len(loss),
            "win_rate_pct":    round(len(wins) / total * 100, 2) if total > 0 else 0.0,
            "profit_factor":   pf,
            "expectancy_r":    round(exp_r, 4),
            "net_pnl_usd":     round(net_pnl, 2),
            "net_return_pct":  round(net_pnl / 10000.0 * 100, 4),  # $10k per-ticker basis
            "gross_wins_usd":  round(gross_w, 2),
            "gross_losses_usd": round(gross_l, 2),
            "avg_win_pct":     round(float(g.loc[g["pnl_dollar"]>0,  "pnl_pct"].mean()), 4) if len(wins)>0 else 0.0,
            "avg_loss_pct":    round(float(g.loc[g["pnl_dollar"]<=0, "pnl_pct"].mean()), 4) if len(loss)>0 else 0.0,
            "best_trade_pct":  round(float(g["pnl_pct"].max()), 4) if total > 0 else 0.0,
            "worst_trade_pct": round(float(g["pnl_pct"].min()), 4) if total > 0 else 0.0,
            "avg_hold_days":   round(float(g["hold_days"].mean()), 1) if total > 0 else 0.0,
            "stage6_entries":  int((g["signal_type"] == "stage6_entry").sum()),
            "stage7_entries":  int((g["signal_type"] == "stage7_entry").sum()),
        })

    return pd.DataFrame(rows).sort_values("expectancy_r", ascending=False).reset_index(drop=True)


def _write_investor_excel(
    df_b1_ticker:  pd.DataFrame,
    df_b2_ticker:  pd.DataFrame,
    df_combined:   pd.DataFrame,
    metrics_b1:    dict,
    metrics_b2:    dict,
    metrics_all:   dict,
    windows:       list,
    source_run:    str,
    out_path:      Path,
) -> None:
    """Write the multi-sheet investor-ready Excel report."""
    try:
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("  [WARNING] openpyxl not installed — Excel export skipped.")
        return

    # ── Style definitions ─────────────────────────────────────────────────────
    H_FILL   = PatternFill("solid", fgColor="1F3864")
    H_FONT   = Font(color="FFFFFF", bold=True, size=10)
    TITLE_F  = Font(bold=True, size=13, color="1F3864")
    SUB_F    = Font(bold=True, size=11, color="2E4057")
    BOLD_F   = Font(bold=True)
    CENTER_A = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT_A   = Alignment(horizontal="left",   vertical="center")
    RIGHT_A  = Alignment(horizontal="right",  vertical="center")
    thin     = Side(style="thin", color="CCCCCC")
    border   = Border(left=thin, right=thin, top=thin, bottom=thin)

    B1_FILL  = PatternFill("solid", fgColor="E3F2FD")   # light blue — Batch 1
    B2_FILL  = PatternFill("solid", fgColor="E8F5E9")   # light green — Batch 2
    COMB_FILL = PatternFill("solid", fgColor="FFF9C4")  # light yellow — combined

    TICKER_COLS = [
        "ticker", "sector", "batch_name", "total_trades", "win_rate_pct",
        "profit_factor", "expectancy_r", "net_return_pct", "net_pnl_usd",
        "avg_win_pct", "avg_loss_pct", "best_trade_pct", "worst_trade_pct",
        "avg_hold_days", "stage7_entries",
        "gross_wins_usd", "gross_losses_usd",
    ]

    COL_LABELS = {
        "ticker": "Ticker", "sector": "Sector", "batch_name": "Batch",
        "total_trades": "Trades", "win_rate_pct": "Win %",
        "profit_factor": "Profit Factor", "expectancy_r": "Expectancy R",
        "net_return_pct": "Net Return %", "net_pnl_usd": "Net PnL ($)",
        "avg_win_pct": "Avg Win %", "avg_loss_pct": "Avg Loss %",
        "best_trade_pct": "Best Trade %", "worst_trade_pct": "Worst Trade %",
        "avg_hold_days": "Avg Hold Days", "stage7_entries": "Stage 7 Entries",
        "gross_wins_usd": "Gross Wins ($)", "gross_losses_usd": "Gross Losses ($)",
    }

    def _write_df_sheet(ws, df: pd.DataFrame, cols: list, fill, title: str):
        """Write a DataFrame sheet with consistent professional formatting."""
        ws.cell(row=1, column=1, value=title).font = TITLE_F
        ws.row_dimensions[1].height = 24
        ws.row_dimensions[2].height = 30

        for c_idx, col in enumerate(cols, 1):
            cell = ws.cell(row=2, column=c_idx,
                           value=COL_LABELS.get(col, col))
            cell.fill = H_FILL; cell.font = H_FONT
            cell.alignment = CENTER_A; cell.border = border

        for r_idx, (_, row) in enumerate(df[cols].iterrows(), 3):
            for c_idx, col in enumerate(cols, 1):
                val = row[col]
                if hasattr(val, "item"): val = val.item()
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.fill   = fill
                cell.border = border
                cell.alignment = RIGHT_A if col not in ["ticker","sector","batch_name"] else LEFT_A

        for c_idx, col in enumerate(cols, 1):
            max_len = max(
                len(COL_LABELS.get(col, col)),
                df[col].astype(str).str.len().max() if len(df) > 0 else 0
            )
            ws.column_dimensions[get_column_letter(c_idx)].width = min(max_len + 3, 28)
        ws.freeze_panes = ws.cell(row=3, column=1)

    def _write_summary_sheet(ws, metrics_list: list, labels: list, fills: list):
        """Write a side-by-side batch summary sheet."""
        ws.cell(row=1, column=1, value="Performance Summary").font = TITLE_F
        ws.cell(row=2, column=1,
                value=f"Source: {source_run}   |   Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}").font = SUB_F

        # Column headers (one per batch)
        ws.cell(row=4, column=1, value="Metric").font = BOLD_F
        for col_idx, (label, fill) in enumerate(zip(labels, fills), 2):
            cell = ws.cell(row=4, column=col_idx, value=label)
            cell.fill = H_FILL; cell.font = H_FONT; cell.alignment = CENTER_A

        metric_rows = [
            ("TRADE STATISTICS", None),
            ("Total Trades",      "total_trades"),
            ("Winning Trades",    "winning_trades"),
            ("Losing Trades",     "losing_trades"),
            ("Win Rate %",        "win_rate_pct"),
            ("Unique Tickers",    "unique_tickers"),
            ("", None),
            ("PERFORMANCE", None),
            ("Profit Factor",     "profit_factor"),
            ("Expectancy R",      "expectancy_r"),
            ("Gross Wins ($)",    "gross_wins_usd"),
            ("Gross Losses ($)",  "gross_losses_usd"),
            ("Net PnL ($)",       "net_pnl_usd"),
            ("", None),
            ("TRADE QUALITY", None),
            ("Avg Win %",         "avg_win_pct"),
            ("Avg Loss %",        "avg_loss_pct"),
            ("Best Trade %",      "best_trade_pct"),
            ("Worst Trade %",     "worst_trade_pct"),
            ("Avg Hold Days",     "avg_hold_days"),
            ("", None),
            ("DATE RANGE", None),
            ("First Entry Date",  "entry_date_first"),
            ("Last Entry Date",   "entry_date_last"),
            ("Last Exit Date",    "exit_date_last"),
            ("", None),
            ("ENTRY STAGES", None),
            ("Stage 6 Entries",   "stage6_entries"),
            ("Stage 7 Entries",   "stage7_entries"),
        ]

        for r_off, (label, key) in enumerate(metric_rows, 5):
            is_section = (key is None and label != "")
            cell_label = ws.cell(row=r_off, column=1, value=label)
            if is_section:
                cell_label.font = Font(bold=True, color="1F3864", size=10)
                cell_label.fill = PatternFill("solid", fgColor="E8EAF6")
                for c in range(1, len(labels) + 2):
                    ws.cell(row=r_off, column=c).fill = PatternFill("solid", fgColor="E8EAF6")

            for col_idx, (metrics, fill) in enumerate(zip(metrics_list, fills), 2):
                if key is None:
                    continue
                val = metrics.get(key, "—")
                if isinstance(val, float):
                    val = round(val, 4)
                if hasattr(val, "item"): val = val.item()
                cell = ws.cell(row=r_off, column=col_idx, value=val)
                cell.fill = fill if not is_section else PatternFill("solid", fgColor="E8EAF6")
                cell.alignment = RIGHT_A
                cell.border = border

        ws.column_dimensions["A"].width = 22
        for col_idx in range(2, len(labels) + 2):
            ws.column_dimensions[get_column_letter(col_idx)].width = 20
        ws.freeze_panes = ws.cell(row=5, column=1)

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        # ── Sheet 1: Performance Summary ─────────────────────────────────────
        writer.book.create_sheet("Performance Summary")
        ws_sum = writer.book["Performance Summary"]
        _write_summary_sheet(
            ws_sum,
            metrics_list=[metrics_b1, metrics_b2, metrics_all],
            labels=[
                windows[0]["name"] if len(windows) > 0 else "Batch 1",
                windows[1]["name"] if len(windows) > 1 else "Batch 2",
                "Combined (All Batches)",
            ],
            fills=[B1_FILL, B2_FILL, COMB_FILL],
        )

        # ── Sheet 2: Batch 1 Ticker Detail ───────────────────────────────────
        if not df_b1_ticker.empty:
            df_b1_ticker.to_excel(writer, index=False, sheet_name="Batch 1 — Tickers")
            _write_df_sheet(
                writer.sheets["Batch 1 — Tickers"],
                df_b1_ticker,
                [c for c in TICKER_COLS if c in df_b1_ticker.columns],
                B1_FILL,
                f"{windows[0]['name'] if len(windows)>0 else 'Batch 1'} — Per-Ticker Performance"
            )

        # ── Sheet 3: Batch 2 Ticker Detail ───────────────────────────────────
        if not df_b2_ticker.empty:
            df_b2_ticker.to_excel(writer, index=False, sheet_name="Batch 2 — Tickers")
            _write_df_sheet(
                writer.sheets["Batch 2 — Tickers"],
                df_b2_ticker,
                [c for c in TICKER_COLS if c in df_b2_ticker.columns],
                B2_FILL,
                f"{windows[1]['name'] if len(windows)>1 else 'Batch 2'} — Per-Ticker Performance"
            )

        # ── Sheet 4: Combined Ticker Summary ─────────────────────────────────
        if not df_combined.empty:
            df_combined.to_excel(writer, index=False, sheet_name="Combined — All Tickers")
            _write_df_sheet(
                writer.sheets["Combined — All Tickers"],
                df_combined,
                [c for c in TICKER_COLS if c in df_combined.columns and c != "batch_name"],
                COMB_FILL,
                "Combined — Per-Ticker Performance Across All Batches"
            )

        # Move Performance Summary to first tab
        wb = writer.book
        wb.move_sheet("Performance Summary", offset=-len(wb.sheetnames) + 1)

    print(f"  investor_report.xlsx  : {out_path}")


def main() -> None:
    print("=" * 70)
    print("09E — TIME-BATCHED ANALYSIS & INVESTOR REPORT")
    print("=" * 70)
    ts_start = datetime.now()

    # ── Load config ───────────────────────────────────────────────────────────
    cfg       = _load_cfg(BACKTEST_CFG)
    batch_cfg = cfg.get("batches", {})

    source_run_tag       = str(batch_cfg.get("source_run_tag", ""))
    use_filtered_tickers = bool(batch_cfg.get("use_filtered_tickers", True))
    windows              = batch_cfg.get("windows", [])

    if not source_run_tag:
        print("[ERROR] batches.source_run_tag not set in config/backtest.yaml")
        sys.exit(1)
    if not windows:
        print("[ERROR] batches.windows not set in config/backtest.yaml")
        sys.exit(1)

    print(f"Source run tag    : {source_run_tag}")
    print(f"Use filter list   : {use_filtered_tickers}")
    print(f"Batch windows     : {len(windows)}")
    for w in windows:
        print(f"  {w['name']}: {w['start']} → {w['end']}")

    # ── Load trades ───────────────────────────────────────────────────────────
    trades_path = BACKTESTS_DIR / source_run_tag / "universe" / "trades_all.parquet"
    if not trades_path.exists():
        print(f"\n[ERROR] trades_all.parquet not found: {trades_path}")
        sys.exit(1)

    trades = pd.read_parquet(trades_path)
    trades["entry_date"] = pd.to_datetime(trades["entry_date"])
    trades["exit_date"]  = pd.to_datetime(trades["exit_date"])
    print(f"\nLoaded trades     : {len(trades):,} trades from {trades['ticker'].nunique():,} tickers")

    # ── Sector enrichment ─────────────────────────────────────────────────────
    # trades_all.parquet already contains spider_id from 09A gate join.
    # Only merge memberships if spider_id column is missing entirely.
    if "spider_id" not in trades.columns:
        if MEMBERSHIPS.exists():
            mem = pd.read_csv(MEMBERSHIPS)[["ticker", "spider_id"]].drop_duplicates("ticker")
            trades = trades.merge(mem, on="ticker", how="left")
        else:
            trades["spider_id"] = None

    trades["sector"] = trades["spider_id"].apply(_sector_from_spider_id)

    # ── Apply ticker filter if requested ─────────────────────────────────────
    filtered_path = FILTER_DIR / "filtered_tickers.csv"
    if use_filtered_tickers:
        if not filtered_path.exists():
            print(f"\n[ERROR] filtered_tickers.csv not found: {filtered_path}")
            print("  Run 09D first to generate the filter list.")
            sys.exit(1)
        ft            = pd.read_csv(filtered_path)
        passing_set   = set(ft["ticker"].tolist())
        before        = len(trades)
        trades        = trades[trades["ticker"].isin(passing_set)].copy()
        print(f"Ticker filter     : {len(trades):,} trades kept ({before - len(trades):,} removed — not in passing list)")
        print(f"Unique tickers    : {trades['ticker'].nunique():,} (from {len(passing_set):,} passing tickers)")
    else:
        print("Ticker filter     : disabled (using full universe)")

    # ── First signal date — start of analysis ────────────────────────────────
    first_signal_date = trades["entry_date"].min()
    last_signal_date  = trades["entry_date"].max()
    print(f"\nEntry date range  : {first_signal_date.date()} → {last_signal_date.date()}")

    # ── Split into batch windows (entry-date attribution) ─────────────────────
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)

    batch_trades_list   = []
    batch_metrics_list  = []
    batch_ticker_frames = []

    for i, window in enumerate(windows, 1):
        w_start = pd.Timestamp(window["start"])
        w_end   = pd.Timestamp(window["end"])
        label   = f"B{i}"

        batch = trades[
            (trades["entry_date"] >= w_start) &
            (trades["entry_date"] <= w_end)
        ].copy()
        batch["batch"]      = label
        batch["batch_name"] = window["name"]

        print(f"\n  {window['name']}:")
        print(f"    Trades          : {len(batch):,}")
        print(f"    Unique tickers  : {batch['ticker'].nunique():,}")

        if batch.empty:
            print(f"    [WARNING] No trades in this window.")
            continue

        # Per-batch aggregate metrics
        metrics = compute_batch_metrics(batch, window["name"])
        metrics["batch_label"]       = label
        metrics["batch_start"]       = str(w_start.date())
        metrics["batch_end"]         = str(w_end.date())
        metrics["first_entry_date"]  = str(batch["entry_date"].min().date())
        metrics["last_entry_date"]   = str(batch["entry_date"].max().date())

        print(f"    Win rate        : {metrics['win_rate_pct']:.1f}%")
        print(f"    Profit factor   : {metrics['profit_factor']}")
        print(f"    Expectancy R    : {metrics['expectancy_r']:.4f}")
        print(f"    Net PnL ($)     : ${metrics['net_pnl_usd']:,.2f}")

        batch_trades_list.append(batch)
        batch_metrics_list.append(metrics)

        # Per-ticker metrics in this batch
        df_ticker = compute_ticker_metrics_in_batch(batch, window["name"], label)
        batch_ticker_frames.append(df_ticker)

        # Save batch parquet
        batch_parquet = REPORTS_DIR / f"batch_{i}_trades.parquet"
        batch.to_parquet(batch_parquet, index=False)

        # Save batch ticker summary CSV
        ticker_csv = REPORTS_DIR / f"batch_{i}_ticker_summary.csv"
        df_ticker.to_csv(ticker_csv, index=False)
        print(f"    Saved parquet   : {batch_parquet.name}")
        print(f"    Saved CSV       : {ticker_csv.name}")

    if not batch_trades_list:
        print("\n[ERROR] No trades in any batch window. Check batch date ranges.")
        sys.exit(1)

    # ── Combined ticker metrics (across all batches) ──────────────────────────
    all_trades = pd.concat(batch_trades_list, ignore_index=True)
    metrics_all = compute_batch_metrics(all_trades, "Combined — All Batches")
    metrics_all["batch_label"]      = "ALL"
    metrics_all["first_entry_date"] = str(all_trades["entry_date"].min().date())
    metrics_all["last_entry_date"]  = str(all_trades["entry_date"].max().date())

    # Aggregate per-ticker across ALL batches (sum trades, recompute metrics)
    combined_rows = []
    for ticker, grp in all_trades.groupby("ticker"):
        g = grp.copy()
        total = len(g)
        wins  = g[g["pnl_dollar"] > 0]
        loss  = g[g["pnl_dollar"] <= 0]
        gross_w = float(wins["pnl_dollar"].sum())       if len(wins) > 0 else 0.0
        gross_l = float(loss["pnl_dollar"].abs().sum()) if len(loss) > 0 else 0.0
        net_pnl = float(g["pnl_dollar"].sum())
        pf      = round(gross_w / gross_l, 4)           if gross_l > 0 else float("inf")
        exp_r   = float(g["pnl_r"].mean())

        combined_rows.append({
            "ticker":          ticker,
            "sector":          g["sector"].iloc[0]    if "sector" in g.columns else "Unknown",
            "spider_id":       g["spider_id"].iloc[0] if "spider_id" in g.columns else None,
            "total_trades":    total,
            "winning_trades":  len(wins),
            "losing_trades":   len(loss),
            "win_rate_pct":    round(len(wins) / total * 100, 2) if total > 0 else 0.0,
            "profit_factor":   pf,
            "expectancy_r":    round(exp_r, 4),
            "net_pnl_usd":     round(net_pnl, 2),
            "net_return_pct":  round(net_pnl / 10000.0 * 100, 4),
            "gross_wins_usd":  round(gross_w, 2),
            "gross_losses_usd": round(gross_l, 2),
            "avg_win_pct":     round(float(g.loc[g["pnl_dollar"]>0,  "pnl_pct"].mean()), 4) if len(wins)>0 else 0.0,
            "avg_loss_pct":    round(float(g.loc[g["pnl_dollar"]<=0, "pnl_pct"].mean()), 4) if len(loss)>0 else 0.0,
            "best_trade_pct":  round(float(g["pnl_pct"].max()), 4),
            "worst_trade_pct": round(float(g["pnl_pct"].min()), 4),
            "avg_hold_days":   round(float(g["hold_days"].mean()), 1),
            "stage6_entries":  int((g["signal_type"] == "stage6_entry").sum()),
            "stage7_entries":  int((g["signal_type"] == "stage7_entry").sum()),
        })

    df_combined = pd.DataFrame(combined_rows).sort_values("expectancy_r", ascending=False)
    df_combined.to_csv(REPORTS_DIR / "combined_ticker_summary.csv", index=False)

    # ── Save JSON reports ─────────────────────────────────────────────────────
    elapsed = (datetime.now() - ts_start).total_seconds()
    full_report = {
        "generated_at": datetime.now().isoformat(),
        "source_run_tag": source_run_tag,
        "elapsed_seconds": round(elapsed, 1),
        "config_snapshot": cfg,  # full backtest.yaml at time of run
        "first_signal_date": str(first_signal_date.date()),
        "last_signal_date":  str(last_signal_date.date()),
        "batches":         batch_metrics_list,
        "combined":        metrics_all,
    }
    with (REPORTS_DIR / "batch_report.json").open("w", encoding="utf-8") as f:
        json.dump(full_report, f, indent=2, default=str)

    # ── Write investor Excel ──────────────────────────────────────────────────
    excel_path = REPORTS_DIR / "investor_report.xlsx"
    b1_tickers = batch_ticker_frames[0] if len(batch_ticker_frames) > 0 else pd.DataFrame()
    b2_tickers = batch_ticker_frames[1] if len(batch_ticker_frames) > 1 else pd.DataFrame()
    m1 = batch_metrics_list[0] if len(batch_metrics_list) > 0 else {}
    m2 = batch_metrics_list[1] if len(batch_metrics_list) > 1 else {}

    _write_investor_excel(
        df_b1_ticker  = b1_tickers,
        df_b2_ticker  = b2_tickers,
        df_combined   = df_combined,
        metrics_b1    = m1,
        metrics_b2    = m2,
        metrics_all   = metrics_all,
        windows       = windows,
        source_run    = source_run_tag,
        out_path      = excel_path,
    )

    # ── Console summary ───────────────────────────────────────────────────────
    print(f"\n{'═' * 55}")
    print(f"  COMBINED SUMMARY (all batches, filtered universe)")
    print(f"{'═' * 55}")
    print(f"  Total trades      : {metrics_all['total_trades']:>8,}")
    print(f"  Unique tickers    : {metrics_all['unique_tickers']:>8,}")
    print(f"  Win rate          : {metrics_all['win_rate_pct']:>8.1f}%")
    print(f"  Profit factor     : {str(metrics_all['profit_factor']):>8}")
    print(f"  Expectancy R      : {metrics_all['expectancy_r']:>8.4f}")
    print(f"  Net PnL ($)       : ${metrics_all['net_pnl_usd']:>12,.2f}")
    print(f"{'═' * 55}")

    print(f"\nOutputs written to  : {REPORTS_DIR}")
    print(f"  investor_report.xlsx")
    print(f"  batch_1_trades.parquet / batch_2_trades.parquet")
    print(f"  batch_1_ticker_summary.csv / batch_2_ticker_summary.csv")
    print(f"  combined_ticker_summary.csv")
    print(f"  batch_report.json")
    print(f"\nElapsed             : {elapsed:.1f}s")
    print("=" * 70)
    print("09E COMPLETE — open output/reports/batched/investor_report.xlsx")
    print("=" * 70)


if __name__ == "__main__":
    main()
