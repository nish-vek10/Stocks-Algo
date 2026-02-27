# Path: research/experiments/09D_filter_enrich_report.py
"""
ALGO-STOCKS Phase 09D — Universe Filter & Enriched Report

PURPOSE
-------
1. Reads the per-ticker summary from a 09B universe run
2. Enriches every ticker with sector information (from spider memberships)
3. Applies configurable filters to identify tickers with genuine edge
4. Writes investor-ready Excel and clean CSVs for further analysis

WHY THIS EXISTS
---------------
The full universe of 2,582 tickers includes companies where the dislocation
→ mean-reversion thesis does not apply: permanently declining businesses,
commodity-driven names, and tickers with too few signals to be statistically
meaningful. This script separates the signal from the noise.

FILTER LOGIC
------------
A ticker PASSES if ALL of the following are true:
  - total_trades >= min_trades
  - profit_factor >= min_profit_factor
  - expectancy_r  >= min_expectancy_r
  - max_drawdown_pct >= max_drawdown_pct (less negative than threshold)
  - win_rate_pct  >= min_win_rate_pct

Thresholds are all configurable in config/backtest.yaml under 'filter:'.

OUTPUTS
-------
output/reports/universe_filter/
    summary_enriched.xlsx         — full universe, all metrics + sector, Excel
    filtered_tickers.csv          — passing tickers only (for 09E input)
    rejected_tickers.csv          — failing tickers with rejection reason
    filter_report.json            — summary statistics of filter run

RUN FROM PROJECT ROOT:
  python research/experiments/09D_filter_enrich_report.py
"""
from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path
import sys

import pandas as pd
import yaml

# ── Project root resolution ───────────────────────────────────────────────────
ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT))

# ══════════════════════════════════════════════════════════════════════════════
# CONFIGURABLE PATHS
# ══════════════════════════════════════════════════════════════════════════════
BACKTEST_CFG  = ROOT / "config" / "backtest.yaml"
BACKTESTS_DIR = ROOT / "output" / "backtests"
MEMBERSHIPS   = ROOT / "data" / "metadata" / "spiders" / "spider_memberships.csv"
REPORTS_DIR   = ROOT / "output" / "reports" / "universe_filter"
# ══════════════════════════════════════════════════════════════════════════════


def _load_cfg(path: Path) -> dict:
    with path.open("r", encoding="utf-8") as f:
        return yaml.safe_load(f) or {}


def _sector_from_spider_id(spider_id: str) -> str:
    """
    Convert spider_id to readable sector name.
    e.g. 'SECTOR_TECHNOLOGY' → 'Technology'
         'SECTOR_HEALTH_CARE' → 'Health Care'
    """
    if not isinstance(spider_id, str):
        return "Unknown"
    name = spider_id.replace("SECTOR_", "").replace("_", " ").title()
    return name


def _build_rejection_reason(row: pd.Series, thresholds: dict) -> str:
    """Return a human-readable string explaining why a ticker was rejected."""
    reasons = []
    if row["total_trades"] < thresholds["min_trades"]:
        reasons.append(f"too_few_trades({int(row['total_trades'])})")
    if row["profit_factor"] < thresholds["min_profit_factor"]:
        reasons.append(f"low_profit_factor({row['profit_factor']:.2f})")
    if row["expectancy_r"] < thresholds["min_expectancy_r"]:
        reasons.append(f"negative_expectancy({row['expectancy_r']:.3f})")
    if row["max_drawdown_pct"] < thresholds["max_drawdown_pct"]:
        reasons.append(f"excess_drawdown({row['max_drawdown_pct']:.1f}%)")
    if row["win_rate_pct"] < thresholds["min_win_rate_pct"]:
        reasons.append(f"low_win_rate({row['win_rate_pct']:.1f}%)")
    return "; ".join(reasons) if reasons else "passed"


def _write_excel(
    df_all:      pd.DataFrame,
    df_pass:     pd.DataFrame,
    df_fail:     pd.DataFrame,
    out_path:    Path,
    thresholds:  dict,
    source_run:  str,
) -> None:
    """
    Write a multi-sheet investor-ready Excel workbook.
    Sheets: Overview | Passing Tickers | Rejected Tickers | Filter Settings
    """
    try:
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.formatting.rule import CellIsRule
    except ImportError:
        print("  [WARNING] openpyxl not installed — Excel export skipped.")
        print("  Install with: pip install openpyxl --break-system-packages")
        return

    # ── Colour palette ────────────────────────────────────────────────────────
    HEADER_FILL  = PatternFill("solid", fgColor="1F3864")   # dark navy
    HEADER_FONT  = Font(color="FFFFFF", bold=True, size=10)
    PASS_FILL    = PatternFill("solid", fgColor="E8F5E9")   # light green
    FAIL_FILL    = PatternFill("solid", fgColor="FFEBEE")   # light red
    NEUTRAL_FILL = PatternFill("solid", fgColor="F5F5F5")   # light grey
    BOLD         = Font(bold=True)
    CENTER       = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT         = Alignment(horizontal="left",   vertical="center")

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Columns to display per sheet
    COLS_FULL = [
        "ticker", "sector", "spider_id", "total_trades", "win_rate_pct",
        "profit_factor", "expectancy_r", "net_return_pct", "max_drawdown_pct",
        "sharpe_ratio", "sortino_ratio", "calmar_ratio",
        "avg_win_pct", "avg_loss_pct", "best_trade_pct", "worst_trade_pct",
        "avg_hold_days", "stage6_entries", "stage7_entries",
        "gross_wins_usd", "gross_losses_usd", "net_pnl_usd",
        "filter_status", "rejection_reason",
    ]

    COLS_PASS = [c for c in COLS_FULL if c not in ["filter_status", "rejection_reason"]]
    COLS_FAIL = [
        "ticker", "sector", "total_trades", "win_rate_pct", "profit_factor",
        "expectancy_r", "net_return_pct", "max_drawdown_pct", "rejection_reason",
    ]

    COL_HEADERS = {
        "ticker": "Ticker", "sector": "Sector", "spider_id": "Spider ID",
        "total_trades": "Total Trades", "win_rate_pct": "Win Rate %",
        "profit_factor": "Profit Factor", "expectancy_r": "Expectancy R",
        "net_return_pct": "Net Return %", "max_drawdown_pct": "Max DD %",
        "sharpe_ratio": "Sharpe", "sortino_ratio": "Sortino",
        "calmar_ratio": "Calmar", "avg_win_pct": "Avg Win %",
        "avg_loss_pct": "Avg Loss %", "best_trade_pct": "Best Trade %",
        "worst_trade_pct": "Worst Trade %", "avg_hold_days": "Avg Hold Days",
        "stage6_entries": "Stage 6", "stage7_entries": "Stage 7",
        "gross_wins_usd": "Gross Wins $", "gross_losses_usd": "Gross Losses $",
        "net_pnl_usd": "Net PnL $", "filter_status": "Status",
        "rejection_reason": "Rejection Reason",
    }

    def _write_sheet(ws, df: pd.DataFrame, cols: list, row_fill=None, title: str = ""):
        """Write a DataFrame to a worksheet with professional formatting."""
        # Title row
        if title:
            ws.row_dimensions[1].height = 22
            cell = ws.cell(row=1, column=1, value=title)
            cell.font = Font(bold=True, size=12, color="1F3864")
            cell.alignment = LEFT

        header_row = 2 if title else 1

        # Header row
        ws.row_dimensions[header_row].height = 32
        for col_idx, col in enumerate(cols, 1):
            cell = ws.cell(row=header_row, column=col_idx,
                           value=COL_HEADERS.get(col, col))
            cell.fill      = HEADER_FILL
            cell.font      = HEADER_FONT
            cell.alignment = CENTER
            cell.border    = border

        # Data rows
        for r_idx, (_, row) in enumerate(df[cols].iterrows(), header_row + 1):
            fill = row_fill if row_fill else NEUTRAL_FILL
            for c_idx, col in enumerate(cols, 1):
                val = row[col]
                # Convert numpy types to native Python for Excel compatibility
                if hasattr(val, "item"):
                    val = val.item()
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.fill      = fill
                cell.border    = border
                cell.alignment = LEFT
                # Right-align numeric columns
                if col not in ["ticker", "sector", "spider_id",
                                "filter_status", "rejection_reason"]:
                    cell.alignment = Alignment(horizontal="right", vertical="center")

        # Auto-size columns
        for col_idx, col in enumerate(cols, 1):
            col_letter = get_column_letter(col_idx)
            max_len = max(
                len(COL_HEADERS.get(col, col)),
                df[col].astype(str).str.len().max() if len(df) > 0 else 0
            )
            ws.column_dimensions[col_letter].width = min(max_len + 3, 30)

        # Freeze header
        ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        # ── Sheet 1: Full Universe ────────────────────────────────────────────
        df_all_export = df_all.reindex(
            columns=[c for c in COLS_FULL if c in df_all.columns]
        ).sort_values("net_return_pct", ascending=False)
        df_all_export.to_excel(writer, index=False, sheet_name="Full Universe")
        ws1 = writer.sheets["Full Universe"]
        _write_sheet(
            ws1, df_all_export,
            [c for c in COLS_FULL if c in df_all_export.columns],
            title=f"Full Universe — {source_run} ({len(df_all_export)} tickers)"
        )

        # ── Sheet 2: Passing Tickers ──────────────────────────────────────────
        df_pass_export = df_pass.reindex(
            columns=[c for c in COLS_PASS if c in df_pass.columns]
        ).sort_values("expectancy_r", ascending=False)
        df_pass_export.to_excel(writer, index=False, sheet_name="Passing Tickers")
        ws2 = writer.sheets["Passing Tickers"]
        _write_sheet(
            ws2, df_pass_export,
            [c for c in COLS_PASS if c in df_pass_export.columns],
            row_fill=PASS_FILL,
            title=f"Passing Tickers ({len(df_pass_export)} tickers — strategy edge confirmed)"
        )

        # ── Sheet 3: Rejected Tickers ─────────────────────────────────────────
        df_fail_export = df_fail.reindex(
            columns=[c for c in COLS_FAIL if c in df_fail.columns]
        ).sort_values("expectancy_r", ascending=True)
        df_fail_export.to_excel(writer, index=False, sheet_name="Rejected Tickers")
        ws3 = writer.sheets["Rejected Tickers"]
        _write_sheet(
            ws3, df_fail_export,
            [c for c in COLS_FAIL if c in df_fail_export.columns],
            row_fill=FAIL_FILL,
            title=f"Rejected Tickers ({len(df_fail_export)} tickers — insufficient edge)"
        )

        # ── Sheet 4: Filter Settings ──────────────────────────────────────────
        ws4 = writer.book.create_sheet("Filter Settings")
        settings = [
            ("Source Run",           source_run),
            ("Generated At",         datetime.now().strftime("%Y-%m-%d %H:%M")),
            ("", ""),
            ("FILTER THRESHOLDS", ""),
            ("Min Trades",           thresholds["min_trades"]),
            ("Min Profit Factor",    thresholds["min_profit_factor"]),
            ("Min Expectancy R",     thresholds["min_expectancy_r"]),
            ("Max Drawdown %",       thresholds["max_drawdown_pct"]),
            ("Min Win Rate %",       thresholds["min_win_rate_pct"]),
            ("", ""),
            ("UNIVERSE SUMMARY", ""),
            ("Total Tickers",        len(df_all)),
            ("Passing Tickers",      len(df_pass)),
            ("Rejected Tickers",     len(df_fail)),
            ("Pass Rate",            f"{len(df_pass)/len(df_all)*100:.1f}%"),
        ]
        for r_idx, (label, value) in enumerate(settings, 1):
            ws4.cell(row=r_idx, column=1, value=label).font = BOLD
            ws4.cell(row=r_idx, column=2, value=value)
        ws4.column_dimensions["A"].width = 25
        ws4.column_dimensions["B"].width = 40

    print(f"  Excel written     : {out_path}")


def main() -> None:
    print("=" * 70)
    print("09D — UNIVERSE FILTER & ENRICHED REPORT")
    print("=" * 70)
    ts_start = datetime.now()

    # ── Load config ───────────────────────────────────────────────────────────
    cfg         = _load_cfg(BACKTEST_CFG)
    run_cfg     = cfg.get("run",    {})
    filter_cfg  = cfg.get("filter", {})
    batch_cfg   = cfg.get("batches", {})

    # Source run: prefer batches.source_run_tag, fall back to run_tag_prefix
    source_run_tag = str(batch_cfg.get("source_run_tag", ""))
    if not source_run_tag:
        print("[ERROR] batches.source_run_tag not set in config/backtest.yaml")
        sys.exit(1)

    # Filter thresholds — all configurable
    thresholds = {
        "min_trades":         int(  filter_cfg.get("min_trades",         5)),
        "min_profit_factor":  float(filter_cfg.get("min_profit_factor",  1.0)),
        "min_expectancy_r":   float(filter_cfg.get("min_expectancy_r",   0.0)),
        "max_drawdown_pct":   float(filter_cfg.get("max_drawdown_pct",  -60.0)),
        "min_win_rate_pct":   float(filter_cfg.get("min_win_rate_pct",   0.0)),
    }

    print(f"Source run tag  : {source_run_tag}")
    print(f"Filter thresholds:")
    for k, v in thresholds.items():
        print(f"  {k:<25s} : {v}")

    # ── Load per-ticker summary ───────────────────────────────────────────────
    summary_path = BACKTESTS_DIR / source_run_tag / "universe" / "summary_by_ticker.csv"
    if not summary_path.exists():
        print(f"\n[ERROR] summary_by_ticker.csv not found: {summary_path}")
        sys.exit(1)

    df = pd.read_csv(summary_path)
    print(f"\nLoaded summary  : {len(df)} tickers from {summary_path.name}")

    # ── Load spider memberships → sector info ─────────────────────────────────
    if MEMBERSHIPS.exists():
        mem = pd.read_csv(MEMBERSHIPS)[["ticker", "spider_id"]].drop_duplicates("ticker")
        df  = df.merge(mem, on="ticker", how="left")
        print(f"Joined sectors  : {df['spider_id'].notna().sum()} / {len(df)} tickers matched")
    else:
        df["spider_id"] = None
        print("[WARNING] spider_memberships.csv not found — sector will be null")

    # Convert spider_id to readable sector name
    df["sector"] = df["spider_id"].apply(_sector_from_spider_id)

    # ── Ensure numeric types ──────────────────────────────────────────────────
    numeric_cols = [
        "total_trades", "win_rate_pct", "profit_factor", "expectancy_r",
        "net_return_pct", "max_drawdown_pct", "sharpe_ratio", "sortino_ratio",
        "calmar_ratio", "avg_win_pct", "avg_loss_pct",
        "best_trade_pct", "worst_trade_pct", "avg_hold_days",
        "stage6_entries", "stage7_entries",
        "gross_wins_usd", "gross_losses_usd", "net_pnl_usd",
    ]
    for col in numeric_cols:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0.0)

    # ── Apply filters ─────────────────────────────────────────────────────────
    mask_pass = (
        (df["total_trades"]     >= thresholds["min_trades"])         &
        (df["profit_factor"]    >= thresholds["min_profit_factor"])  &
        (df["expectancy_r"]     >= thresholds["min_expectancy_r"])   &
        (df["max_drawdown_pct"] >= thresholds["max_drawdown_pct"])   &
        (df["win_rate_pct"]     >= thresholds["min_win_rate_pct"])
    )

    df["rejection_reason"] = df.apply(
        lambda row: _build_rejection_reason(row, thresholds), axis=1
    )
    df["filter_status"] = mask_pass.map({True: "PASS", False: "FAIL"})

    df_pass = df[mask_pass].copy()
    df_fail = df[~mask_pass].copy()

    print(f"\nFilter results:")
    print(f"  Total tickers   : {len(df):,}")
    print(f"  Passed          : {len(df_pass):,}  ({len(df_pass)/len(df)*100:.1f}%)")
    print(f"  Rejected        : {len(df_fail):,}  ({len(df_fail)/len(df)*100:.1f}%)")

    # ── Sector breakdown of passing tickers ──────────────────────────────────
    print(f"\nPassing tickers by sector:")
    sector_counts = (
        df_pass.groupby("sector")
        .agg(tickers=("ticker", "count"),
             avg_pf=("profit_factor", "mean"),
             avg_exp_r=("expectancy_r", "mean"))
        .sort_values("tickers", ascending=False)
    )
    for sector, row in sector_counts.iterrows():
        print(f"  {sector:<25s} : {int(row['tickers']):>4} tickers  "
              f"avg PF={row['avg_pf']:.2f}  avg ExpR={row['avg_exp_r']:.3f}")

    # ── Write outputs ─────────────────────────────────────────────────────────
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)

    # Filtered tickers list — used as input by 09E
    filtered_tickers_path = REPORTS_DIR / "filtered_tickers.csv"
    df_pass[["ticker", "sector", "spider_id"]].to_csv(
        filtered_tickers_path, index=False
    )
    print(f"\nOutputs:")
    print(f"  filtered_tickers.csv  : {filtered_tickers_path}")

    # Rejected tickers — full metrics saved for future reference and reporting
    # Sorted by expectancy_r descending so "closest to passing" appear first
    rejected_path = REPORTS_DIR / "rejected_tickers.csv"
    REJECTED_COLS = [
        "ticker", "sector", "spider_id",
        "total_trades", "winning_trades", "losing_trades",
        "win_rate_pct", "profit_factor", "expectancy_r",
        "net_return_pct", "net_pnl_usd",
        "avg_win_pct", "avg_loss_pct",
        "best_trade_pct", "worst_trade_pct",
        "gross_wins_usd", "gross_losses_usd",
        "max_drawdown_pct", "sharpe_ratio", "sortino_ratio", "calmar_ratio",
        "avg_hold_days", "max_hold_days",
        "stage6_entries", "stage7_entries",
        "rejection_reason",
    ]
    rejected_export = df_fail.reindex(
        columns=[c for c in REJECTED_COLS if c in df_fail.columns]
    ).sort_values("expectancy_r", ascending=False)

    rejected_export.to_csv(rejected_path, index=False)
    print(f"  rejected_tickers.csv  : {rejected_path}  ({len(rejected_export)} tickers)")

    # Standalone Excel for rejected tickers — useful for investor reporting
    rejected_excel_path = REPORTS_DIR / "rejected_tickers.xlsx"
    try:
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        FAIL_FILL = PatternFill("solid", fgColor="FFEBEE")  # light red
        HEADER_FILL = PatternFill("solid", fgColor="7F0000")  # dark red
        HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
        CENTER_A = Alignment(horizontal="center", vertical="center", wrap_text=True)
        LEFT_A = Alignment(horizontal="left", vertical="center")
        RIGHT_A = Alignment(horizontal="right", vertical="center")
        thin = Side(style="thin", color="CCCCCC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        COL_LABELS = {
            "ticker": "Ticker", "sector": "Sector", "spider_id": "Spider ID",
            "total_trades": "Total Trades", "winning_trades": "Winning",
            "losing_trades": "Losing", "win_rate_pct": "Win Rate %",
            "profit_factor": "Profit Factor", "expectancy_r": "Expectancy R",
            "net_return_pct": "Net Return %", "net_pnl_usd": "Net PnL ($)",
            "avg_win_pct": "Avg Win %", "avg_loss_pct": "Avg Loss %",
            "best_trade_pct": "Best Trade %", "worst_trade_pct": "Worst Trade %",
            "gross_wins_usd": "Gross Wins ($)", "gross_losses_usd": "Gross Losses ($)",
            "max_drawdown_pct": "Max DD %", "sharpe_ratio": "Sharpe",
            "sortino_ratio": "Sortino", "calmar_ratio": "Calmar",
            "avg_hold_days": "Avg Hold Days", "max_hold_days": "Max Hold Days",
            "stage6_entries": "Stage 6", "stage7_entries": "Stage 7",
            "rejection_reason": "Rejection Reason",
        }

        display_cols = [c for c in REJECTED_COLS if c in rejected_export.columns]

        with pd.ExcelWriter(rejected_excel_path, engine="openpyxl") as writer:
            rejected_export.to_excel(writer, index=False, sheet_name="Rejected Tickers")
            ws = writer.sheets["Rejected Tickers"]

            # Title
            title_cell = ws.cell(row=1, column=1,
                                 value=f"Rejected Tickers — {source_run_tag}  |  "
                                       f"{len(rejected_export)} tickers failed filter  |  "
                                       f"Generated {datetime.now().strftime('%Y-%m-%d %H:%M')}")
            title_cell.font = Font(bold=True, size=11, color="7F0000")
            ws.row_dimensions[1].height = 22
            ws.row_dimensions[2].height = 30

            # Headers
            for c_idx, col in enumerate(display_cols, 1):
                cell = ws.cell(row=2, column=c_idx,
                               value=COL_LABELS.get(col, col))
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.alignment = CENTER_A
                cell.border = border

            # Data rows
            for r_idx, (_, row) in enumerate(rejected_export[display_cols].iterrows(), 3):
                for c_idx, col in enumerate(display_cols, 1):
                    val = row[col]
                    if hasattr(val, "item"): val = val.item()
                    cell = ws.cell(row=r_idx, column=c_idx, value=val)
                    cell.fill = FAIL_FILL
                    cell.border = border
                    cell.alignment = (
                        LEFT_A if col in ["ticker", "sector", "spider_id", "rejection_reason"]
                        else RIGHT_A
                    )

            # Auto-size columns
            for c_idx, col in enumerate(display_cols, 1):
                max_len = max(
                    len(COL_LABELS.get(col, col)),
                    rejected_export[col].astype(str).str.len().max()
                    if len(rejected_export) > 0 else 0
                )
                ws.column_dimensions[
                    get_column_letter(c_idx)
                ].width = min(max_len + 3, 35)

            ws.freeze_panes = ws.cell(row=3, column=1)

        print(f"  rejected_tickers.xlsx : {rejected_excel_path}  ({len(rejected_export)} tickers)")

    except ImportError:
        print("  [WARNING] openpyxl not installed — rejected Excel skipped.")
        print("  Install with: pip install openpyxl --break-system-packages")

    # Full enriched summary CSV
    summary_enriched_path = REPORTS_DIR / "summary_enriched.csv"
    df.sort_values("expectancy_r", ascending=False).to_csv(
        summary_enriched_path, index=False
    )
    print(f"  summary_enriched.csv  : {summary_enriched_path}")

    # Excel workbook
    excel_path = REPORTS_DIR / "universe_filter_report.xlsx"
    _write_excel(df, df_pass, df_fail, excel_path, thresholds, source_run_tag)

    # Filter report JSON — includes full config snapshot for audit trail
    elapsed = (datetime.now() - ts_start).total_seconds()
    report = {
        "generated_at": datetime.now().isoformat(),
        "source_run_tag": source_run_tag,
        "config_snapshot": cfg,  # full backtest.yaml at time of run
        "elapsed_seconds":   round(elapsed, 1),
        "total_tickers":     len(df),
        "passed":            len(df_pass),
        "rejected":          len(df_fail),
        "pass_rate_pct":     round(len(df_pass) / len(df) * 100, 2),
        "thresholds":        thresholds,
        "sector_breakdown":  sector_counts.reset_index().to_dict(orient="records"),
        "rejection_reasons": df_fail["rejection_reason"].value_counts().head(20).to_dict(),
    }
    json_path = REPORTS_DIR / "filter_report.json"
    with json_path.open("w", encoding="utf-8") as f:
        json.dump(report, f, indent=2, default=str)
    print(f"  filter_report.json    : {json_path}")

    print(f"\nElapsed         : {elapsed:.1f}s")
    print("=" * 70)
    print("09D COMPLETE — Run 09E to generate batched investor report")
    print("=" * 70)


if __name__ == "__main__":
    main()
