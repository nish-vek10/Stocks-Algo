# Path: research/experiments/09F_full_universe_portfolio.py
"""
ALGO-STOCKS Phase 09F — Full Universe Virtual Portfolio (Zero-Blocking Baseline)

PURPOSE
-------
Answers the question that 09C structurally cannot answer:
"What does the complete signal set look like as a portfolio with ZERO
 capital competition distortion?"

09C blocks trades when capital runs out. That means every 09C run measures
a capital-selection-biased SUBSET of the strategy — not the strategy itself.
09F fixes this by giving every trade its own dedicated virtual account so all
44,008 trades execute, and then aggregates them into one portfolio equity curve.

METHODOLOGY — Virtual Parallel Accounts
-----------------------------------------
Each trade runs in its own isolated $10,000 virtual account (identical to 09B).
All trades execute in parallel — zero blocking, zero competition.

A single daily portfolio equity curve is built by summing every position's
contribution simultaneously. Correlation is fully captured: on a day in 2022
when 800 positions are open and the market falls, all 800 lose together — the
equity curve drops accordingly. This gives you real portfolio drawdown.

RETURN METRIC HIERARCHY
-----------------------
Three return metrics are reported, each with a different denominator:

1. Ann. Return on Avg Deployed Capital  ← PRIMARY METRIC for investor comparison
   = (net_pnl / avg_deployed_capital / years) × 100
   avg_deployed = avg_open_positions × $10k = capital actually working at any moment
   This is directly comparable to 09C's annualised return.

2. CAGR on Total Virtual Capital  ← conservative, full-base view
   = (final / start)^(1/years) - 1
   denominator = n_trades × $10k = maximum theoretical capital demand

3. Net Return on Total Virtual Capital  ← informational only
   = net_pnl / (n_trades × $10k)
   Not meaningful as a deployment figure — shown for completeness only.

METRIC GROUPS (A–G)
-------------------
A. Returns          — PnL, annualised return, CAGR
B. Risk             — drawdown, duration, recovery, volatility
C. Risk-Adjusted    — Sharpe, Sortino, Calmar, Omega
D. Trade Statistics — win rate, PF, expectancy R, payoff ratio, hold time
E. Consistency      — monthly/quarterly positive rate, best/worst period
F. Position Stats   — avg/max concurrent, deployed capital, utilisation
G. Exit Quality     — exit reason breakdown (Stage9 vs stop vs time)

OUTPUTS
-------
output/reports/full_universe_portfolio/
    full_universe_report.xlsx   ← investor-grade 5-sheet Excel (primary)
    daily_equity.parquet        ← daily equity curve
    monthly_returns.csv         ← year × month return table
    trade_contributions.csv     ← per-trade contribution log
    portfolio_summary.json      ← all metrics + config snapshot (audit trail)

RUN FROM PROJECT ROOT:
  python research/experiments/09F_full_universe_portfolio.py
"""
from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path
import sys

import numpy as np
import pandas as pd
import yaml

# ── Project root ──────────────────────────────────────────────────────────────
ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT))

# ══════════════════════════════════════════════════════════════════════════════
# CONFIGURABLE SETTINGS — edit here, nowhere else
# ══════════════════════════════════════════════════════════════════════════════
RUN_TAG = "filtered"
BACKTEST_CFG  = ROOT / "config" / "backtest.yaml"
BACKTESTS_DIR = ROOT / "output" / "backtests"
FILTER_DIR    = ROOT / "output" / "reports" / "universe_filter"
MEMBERSHIPS   = ROOT / "data" / "metadata" / "spiders" / "spider_memberships.csv"
REPORTS_DIR   = ROOT / "output" / "reports" / "full_universe_portfolio" / RUN_TAG

# Virtual account size per trade — must match 09B account_equity for comparability
VIRTUAL_ACCOUNT_PER_TRADE: float = 10_000.0

# False = full universe (all 44,008 trades) — true honest baseline
# True  = filtered universe only (tickers that passed 09D filter)
USE_FILTERED_TICKERS: bool = True
# ══════════════════════════════════════════════════════════════════════════════


# ── Colour palette — consistent with 09D/09E Excel outputs ───────────────────
NAVY        = "1F3864"
BLUE        = "2E75B6"
LIGHT_BLUE  = "D5E8F0"
L_GREEN     = "E8F5E9"
L_RED       = "FFEBEE"
L_GREY      = "F5F5F5"
MID_GREY    = "CCCCCC"
WHITE       = "FFFFFF"
DARK_GREY   = "404040"
D_GREEN     = "1B5E20"
D_RED       = "B71C1C"
PALE_GREEN  = "C8E6C9"
PALE_RED    = "FFCDD2"
AMBER       = "FFF9C4"


# ─────────────────────────────────────────────────────────────────────────────
# STEP 1 — Load and prepare trades
# ─────────────────────────────────────────────────────────────────────────────

def _load_cfg(path: Path) -> dict:
    with path.open("r", encoding="utf-8") as f:
        return yaml.safe_load(f) or {}


def _sector_from_spider_id(spider_id) -> str:
    if not isinstance(spider_id, str):
        return "Unknown"
    return spider_id.replace("SECTOR_", "").replace("_", " ").title()


# ─────────────────────────────────────────────────────────────────────────────
# STEP 2 — Build daily portfolio equity curve
# ─────────────────────────────────────────────────────────────────────────────

def build_daily_equity_curve(trades: pd.DataFrame) -> pd.DataFrame:
    """
    Build daily portfolio equity as: reference_capital + cumulative_realised_PnL.

    WHY THIS APPROACH
    -----------------
    The naive approach (sum all virtual accounts) grows monotonically — even
    a losing trade returns $9,950 of $10,000 back, so equity never meaningfully
    drops. Every drawdown/Sharpe/Sortino metric computed from that curve is
    meaningless.

    The correct approach:
      1. Fix a reference capital = avg_deployed_capital (capital actually working
         on an average day in the strategy).
      2. Compute the net PnL arriving each day from all trades that exit that day.
      3. equity(t) = reference_capital + sum(daily_pnl[0..t])

    This gives a curve that genuinely fluctuates:
      - Clusters of stop-outs → equity drops below reference → real drawdown
      - Winning streaks → equity rises above reference → real highs
    All downstream metrics (max DD, Sharpe, Sortino, VaR) are now meaningful.

    Reference capital is set to avg_deployed = avg_open_positions × $10k,
    computed from the actual trade data. This represents the capital "at work"
    on an average day and is the right denominator for return comparisons.

    Returns:
        DataFrame: [date, equity, open_positions, daily_pnl,
                    cumulative_pnl, reference_capital, drawdown_pct]
    """
    t = trades.copy()
    t["entry_date"] = pd.to_datetime(t["entry_date"])
    t["exit_date"]  = pd.to_datetime(t["exit_date"])

    # Business-day calendar spanning full history
    all_dates = pd.bdate_range(
        start=t["entry_date"].min(),
        end=t["exit_date"].max(),
    )

    # Pre-aggregate open/close counts by date for O(n) iteration
    entry_counts = t.groupby("entry_date").size().reindex(all_dates, fill_value=0)
    exit_counts  = t.groupby("exit_date").size().reindex(all_dates,  fill_value=0)
    exit_pnls    = (
        t.groupby("exit_date")["pnl_dollar"]
        .sum()
        .reindex(all_dates, fill_value=0.0)
    )

    # First pass — compute avg open positions to set reference capital
    open_now = 0
    open_series = []
    for date in all_dates:
        open_now -= int(exit_counts[date])
        open_now += int(entry_counts[date])
        open_series.append(max(open_now, 0))
    avg_open = float(np.mean(open_series)) if open_series else 1.0
    reference_capital = avg_open * VIRTUAL_ACCOUNT_PER_TRADE

    # Second pass — build equity curve
    rows: list[dict] = []
    cumulative_pnl = 0.0
    open_now       = 0

    for date, open_ct in zip(all_dates, open_series):
        daily_pnl       = float(exit_pnls[date])
        cumulative_pnl += daily_pnl
        open_now        = open_ct
        equity          = reference_capital + cumulative_pnl

        rows.append({
            "date":              date,
            "equity":            equity,
            "open_positions":    open_now,
            "daily_pnl":         daily_pnl,
            "cumulative_pnl":    cumulative_pnl,
            "reference_capital": reference_capital,
            "deployed_capital":  open_now * VIRTUAL_ACCOUNT_PER_TRADE,
        })

    df = pd.DataFrame(rows)
    df["date"] = pd.to_datetime(df["date"])

    # Add drawdown column here for convenience
    peak          = df["equity"].cummax()
    df["drawdown_pct"] = ((df["equity"] - peak) / peak * 100).round(4)
    return df


# ─────────────────────────────────────────────────────────────────────────────
# STEP 3 — Full quant metric suite (Groups A–G)
# ─────────────────────────────────────────────────────────────────────────────

def compute_all_metrics(
    daily_eq:     pd.DataFrame,
    trades:       pd.DataFrame,
    start_equity: float,
) -> dict:
    """
    Compute the full institutional metric suite across seven groups.

    A. Return metrics
    B. Risk metrics
    C. Risk-adjusted ratios
    D. Trade statistics
    E. Consistency (monthly / quarterly)
    F. Position & deployment statistics
    G. Exit quality breakdown
    """
    eq     = daily_eq["equity"].values.astype(float)
    dates  = pd.to_datetime(daily_eq["date"])
    eq_s   = pd.Series(eq, index=dates)

    daily_ret = eq_s.pct_change().dropna()
    n_days    = len(daily_ret)
    years     = n_days / 252.0

    final_equity = float(eq[-1])
    net_pnl      = float(trades["pnl_dollar"].sum())

    # ── A. Returns ────────────────────────────────────────────────────────────

    # Net return on total virtual capital (informational)
    net_return_total_pct = (final_equity - start_equity) / start_equity * 100

    # Return on average deployed capital (most meaningful)
    avg_deployed = float(daily_eq["deployed_capital"].mean())
    ret_on_avg_deployed_pct = (
        net_pnl / avg_deployed * 100 if avg_deployed > 0 else 0.0
    )

    # Annualised return on average deployed — PRIMARY metric
    ann_ret_on_avg_deployed_pct = (
        ret_on_avg_deployed_pct / years if years > 0 else 0.0
    )

    # CAGR on total virtual capital (conservative full-base view)
    cagr_pct = (
        ((final_equity / start_equity) ** (1.0 / years) - 1) * 100
        if years > 0 and start_equity > 0 else 0.0
    )

    # ── B. Risk ───────────────────────────────────────────────────────────────

    running_peak  = eq_s.cummax()
    dd_series_pct = (eq_s - running_peak) / running_peak * 100
    max_dd_pct    = float(dd_series_pct.min())

    # Max drawdown duration — longest consecutive days in drawdown
    in_dd    = (dd_series_pct < 0).astype(int)
    dd_group = (in_dd != in_dd.shift()).cumsum()
    dd_runs  = in_dd.groupby(dd_group).transform("sum")
    max_dd_duration_days = int(dd_runs.max()) if len(dd_runs) > 0 else 0

    # Recovery time from the max drawdown trough to new equity high
    trough_idx  = dd_series_pct.idxmin()
    post_trough = eq_s.loc[trough_idx:]
    peak_at_dd  = float(running_peak.loc[trough_idx])
    recovered   = post_trough[post_trough >= peak_at_dd]
    recovery_days = (
        int((recovered.index[0] - trough_idx).days)
        if len(recovered) > 0 else -1
    )

    # Average drawdown (negative periods only)
    avg_dd_pct = float(
        dd_series_pct[dd_series_pct < 0].mean()
    ) if (dd_series_pct < 0).any() else 0.0

    # Annualised volatility of daily returns
    ann_vol_pct = float(daily_ret.std() * np.sqrt(252) * 100)

    # Pain index — average magnitude of drawdown across ALL days
    pain_index = float(dd_series_pct.abs().mean())

    # Value at Risk (95%) — 5th percentile of daily returns
    var_95 = float(np.percentile(daily_ret.values, 5) * 100)

    # Conditional VaR / Expected Shortfall (95%)
    cvar_95 = float(
        daily_ret[daily_ret <= np.percentile(daily_ret.values, 5)].mean() * 100
    ) if len(daily_ret) > 0 else 0.0

    # ── C. Risk-adjusted ratios ───────────────────────────────────────────────

    mean_ret  = float(daily_ret.mean())
    std_ret   = float(daily_ret.std())
    dn_ret    = daily_ret[daily_ret < 0]
    dn_std    = float(dn_ret.std()) if len(dn_ret) > 1 else 1e-9

    sharpe  = (mean_ret / std_ret  * np.sqrt(252)) if std_ret  > 0 else 0.0
    sortino = (mean_ret / dn_std   * np.sqrt(252)) if dn_std   > 0 else 0.0

    # Calmar uses annualised return on avg deployed vs max drawdown
    calmar = (
        (ann_ret_on_avg_deployed_pct / 100) / abs(max_dd_pct / 100)
        if max_dd_pct < 0 else 0.0
    )

    # Omega ratio — probability-weighted upside / probability-weighted downside
    gains      = daily_ret[daily_ret > 0]
    losses_neg = daily_ret[daily_ret <= 0]
    omega = (
        float(gains.sum() / abs(losses_neg.sum()))
        if len(losses_neg) > 0 and losses_neg.sum() != 0 else float("inf")
    )

    # ── D. Trade statistics ───────────────────────────────────────────────────

    total  = len(trades)
    wins   = trades[trades["pnl_dollar"] > 0]
    losses = trades[trades["pnl_dollar"] <= 0]
    w_cnt  = len(wins)
    l_cnt  = len(losses)

    gross_w = float(wins["pnl_dollar"].sum())          if w_cnt > 0 else 0.0
    gross_l = float(losses["pnl_dollar"].abs().sum())  if l_cnt > 0 else 0.0
    pf      = round(gross_w / gross_l, 4)              if gross_l > 0 else float("inf")
    exp_r   = float(trades["pnl_r"].mean())            if total   > 0 else 0.0
    win_rate_pct = w_cnt / total * 100                 if total   > 0 else 0.0

    avg_win_pct = float(wins["pnl_pct"].median()) if w_cnt > 0 else 0.0
    avg_loss_pct = float(losses["pnl_pct"].median()) if l_cnt > 0 else 0.0

    avg_win_r = float(wins["pnl_r"].mean()) if w_cnt > 0 else 0.0
    avg_loss_r = float(losses["pnl_r"].mean()) if l_cnt > 0 else 0.0
    payoff_ratio = abs(avg_win_r / avg_loss_r) if avg_loss_r != 0 else float("inf")

    best_trade_pct = float(trades["pnl_pct"].max()) if total > 0 else 0.0
    worst_trade_pct = float(trades["pnl_pct"].min()) if total > 0 else 0.0

    # Top 5 / bottom 5 outlier trades — fully transparent, not hidden
    top5_trades = (
        trades.nlargest(5, "pnl_pct")[["ticker", "entry_date", "exit_date", "pnl_pct", "pnl_r", "exit_reason"]]
        .assign(pnl_pct=lambda x: x["pnl_pct"].round(2))
        .to_dict("records")
    ) if total > 0 else []
    bot5_trades = (
        trades.nsmallest(5, "pnl_pct")[["ticker", "entry_date", "exit_date", "pnl_pct", "pnl_r", "exit_reason"]]
        .assign(pnl_pct=lambda x: x["pnl_pct"].round(2))
        .to_dict("records")
    ) if total > 0 else []
    avg_hold_days   = float(trades["hold_days"].mean())     if total > 0 else 0.0
    max_hold_days   = int(trades["hold_days"].max())        if total > 0 else 0

    # Consecutive wins / losses
    pnl_sign = (trades.sort_values("entry_date")["pnl_dollar"] > 0).astype(int)
    grp      = (pnl_sign != pnl_sign.shift()).cumsum()
    run_lens = pnl_sign.groupby(grp).transform("count")
    win_runs  = pnl_sign.groupby(grp).transform("count")[pnl_sign == 1]
    loss_runs = pnl_sign.groupby(grp).transform("count")[pnl_sign == 0]
    max_consec_wins   = int(win_runs.max())  if len(win_runs)  > 0 else 0
    max_consec_losses = int(loss_runs.max()) if len(loss_runs) > 0 else 0

    # ── E. Consistency — monthly / quarterly ─────────────────────────────────

    monthly_eq  = eq_s.resample("ME").last()
    monthly_ret = monthly_eq.pct_change().dropna()
    pos_months  = float((monthly_ret > 0).mean() * 100) if len(monthly_ret) > 0 else 0.0
    best_month  = float(monthly_ret.max() * 100)        if len(monthly_ret) > 0 else 0.0
    worst_month = float(monthly_ret.min() * 100)        if len(monthly_ret) > 0 else 0.0
    avg_month   = float(monthly_ret.mean() * 100)       if len(monthly_ret) > 0 else 0.0

    quarterly_eq  = eq_s.resample("QE").last()
    quarterly_ret = quarterly_eq.pct_change().dropna()
    pos_quarters  = float((quarterly_ret > 0).mean() * 100) if len(quarterly_ret) > 0 else 0.0

    annual_eq  = eq_s.resample("YE").last()
    annual_ret = annual_eq.pct_change().dropna()
    pos_years  = float((annual_ret > 0).mean() * 100) if len(annual_ret) > 0 else 0.0

    # ── F. Position & deployment ──────────────────────────────────────────────

    avg_open     = float(daily_eq["open_positions"].mean())
    max_open     = int(daily_eq["open_positions"].max())
    avg_deployed = float(daily_eq["deployed_capital"].mean())
    peak_deployed= float(daily_eq["deployed_capital"].max())
    avg_util_pct = avg_deployed  / start_equity * 100 if start_equity > 0 else 0.0
    peak_util_pct= peak_deployed / start_equity * 100 if start_equity > 0 else 0.0

    # Stage 6 vs Stage 7 split
    s6 = int((trades["signal_type"] == "stage6_entry").sum()) if "signal_type" in trades.columns else 0
    s7 = int((trades["signal_type"] == "stage7_entry").sum()) if "signal_type" in trades.columns else 0

    # ── G. Exit reasons ───────────────────────────────────────────────────────
    exit_raw = trades["exit_reason"].value_counts()
    exit_reasons = {
        r: {"count": int(c), "pct": round(c / total * 100, 1)}
        for r, c in exit_raw.items()
    }

    return {
        # A. Returns
        "net_pnl_usd":                      round(net_pnl, 2),
        "gross_wins_usd":                   round(gross_w, 2),
        "gross_losses_usd":                 round(gross_l, 2),
        "ann_ret_on_avg_deployed_pct":      round(ann_ret_on_avg_deployed_pct, 2),
        "ret_on_avg_deployed_pct":          round(ret_on_avg_deployed_pct, 2),
        "cagr_total_base_pct":              round(cagr_pct, 2),
        "net_return_total_base_pct":        round(net_return_total_pct, 2),
        "start_equity":                     round(start_equity, 2),
        "final_equity":                     round(final_equity, 2),
        # B. Risk
        "max_drawdown_pct":                 round(max_dd_pct, 2),
        "max_dd_duration_days":             max_dd_duration_days,
        "recovery_days":                    recovery_days,
        "avg_drawdown_pct":                 round(avg_dd_pct, 2),
        "ann_volatility_pct":               round(ann_vol_pct, 2),
        "pain_index":                       round(pain_index, 4),
        "var_95_pct":                       round(var_95, 4),
        "cvar_95_pct":                      round(cvar_95, 4),
        # C. Risk-adjusted
        "sharpe_ratio":                     round(sharpe, 4),
        "sortino_ratio":                    round(sortino, 4),
        "calmar_ratio":                     round(calmar, 4),
        "omega_ratio":                      round(omega, 4),
        # D. Trades
        "total_trades":                     total,
        "winning_trades":                   w_cnt,
        "losing_trades":                    l_cnt,
        "win_rate_pct":                     round(win_rate_pct, 2),
        "profit_factor":                    pf,
        "expectancy_r":                     round(exp_r, 4),
        "payoff_ratio":                     round(payoff_ratio, 4),
        "avg_win_pct":                      round(avg_win_pct, 2),
        "avg_loss_pct":                     round(avg_loss_pct, 2),
        "avg_win_r":                        round(avg_win_r, 4),
        "avg_loss_r":                       round(avg_loss_r, 4),
        "best_trade_pct":                   round(best_trade_pct, 2),
        "worst_trade_pct":                  round(worst_trade_pct, 2),
        "avg_hold_days":                    round(avg_hold_days, 1),
        "max_hold_days":                    max_hold_days,
        "max_consec_wins":                  max_consec_wins,
        "max_consec_losses":                max_consec_losses,
        "stage6_entries":                   s6,
        "stage7_entries":                   s7,
        "top5_trades":                      top5_trades,
        "bot5_trades":                      bot5_trades,
        # E. Consistency
        "pct_positive_months":              round(pos_months, 1),
        "pct_positive_quarters":            round(pos_quarters, 1),
        "pct_positive_years":               round(pos_years, 1),
        "best_month_pct":                   round(best_month, 2),
        "worst_month_pct":                  round(worst_month, 2),
        "avg_month_pct":                    round(avg_month, 2),
        # F. Positions
        "avg_open_positions":               round(avg_open, 1),
        "max_open_positions":               max_open,
        "avg_deployed_capital":             round(avg_deployed, 2),
        "peak_deployed_capital":            round(peak_deployed, 2),
        "avg_utilisation_pct":              round(avg_util_pct, 2),
        "peak_utilisation_pct":             round(peak_util_pct, 2),
        "virtual_acct_per_trade":           VIRTUAL_ACCOUNT_PER_TRADE,
        "total_virtual_capital":            round(start_equity, 2),
        "tickers_included":                 int(trades["ticker"].nunique()),
        "first_entry_date":                 str(trades["entry_date"].min().date()),
        "last_exit_date":                   str(trades["exit_date"].max().date()),
        # G. Exit quality
        "exit_reasons":                     exit_reasons,
    }


# ─────────────────────────────────────────────────────────────────────────────
# STEP 4 — Monthly return calendar table
# ─────────────────────────────────────────────────────────────────────────────

def build_monthly_table(daily_eq: pd.DataFrame) -> pd.DataFrame:
    """Year × month return table (%) for heat-map style presentation."""
    eq      = pd.Series(
        daily_eq["equity"].values,
        index=pd.to_datetime(daily_eq["date"])
    )
    monthly = eq.resample("ME").last().pct_change() * 100

    rows: dict = {}
    for dt, val in monthly.items():
        yr  = dt.year
        mon = dt.strftime("%b")
        rows.setdefault(yr, {})[mon] = round(val, 2)

    months = ["Jan","Feb","Mar","Apr","May","Jun",
              "Jul","Aug","Sep","Oct","Nov","Dec"]
    df = pd.DataFrame(rows).T.reindex(columns=months)
    df.index.name = "Year"

    # Annual return per year
    annual = eq.resample("YE").last().pct_change() * 100
    annual_map = {dt.year: round(v, 2) for dt, v in annual.items()}
    df["Annual %"] = [annual_map.get(yr, float("nan")) for yr in df.index]
    return df


# ─────────────────────────────────────────────────────────────────────────────
# STEP 5 — Sector breakdown table
# ─────────────────────────────────────────────────────────────────────────────

def build_sector_breakdown(trades: pd.DataFrame) -> pd.DataFrame:
    """Full per-sector metric breakdown."""
    if "sector" not in trades.columns:
        return pd.DataFrame()

    rows = []
    for sector, g in trades.groupby("sector"):
        total  = len(g)
        wins   = g[g["pnl_dollar"] > 0]
        losses = g[g["pnl_dollar"] <= 0]
        gw     = float(wins["pnl_dollar"].sum())          if len(wins)   > 0 else 0.0
        gl     = float(losses["pnl_dollar"].abs().sum())  if len(losses) > 0 else 0.0
        rows.append({
            "Sector":          sector,
            "Trades":          total,
            "Win Rate %":      round(len(wins) / total * 100, 1) if total > 0 else 0.0,
            "Profit Factor":   round(gw / gl, 4)                  if gl   > 0 else float("inf"),
            "Expectancy R":    round(float(g["pnl_r"].mean()), 4),
            "Net PnL ($)":     round(float(g["pnl_dollar"].sum()), 2),
            "Avg Win %":       round(float(wins["pnl_pct"].mean()), 2) if len(wins)   > 0 else 0.0,
            "Avg Loss %":      round(float(losses["pnl_pct"].mean()), 2) if len(losses) > 0 else 0.0,
            "Avg Hold Days":   round(float(g["hold_days"].mean()), 1),
        })

    return (
        pd.DataFrame(rows)
        .sort_values("Net PnL ($)", ascending=False)
        .reset_index(drop=True)
    )


# ─────────────────────────────────────────────────────────────────────────────
# STEP 6 — Investor-grade Excel workbook (5 sheets)
# ─────────────────────────────────────────────────────────────────────────────

def write_investor_excel(
    metrics:     dict,
    daily_eq:    pd.DataFrame,
    monthly_tbl: pd.DataFrame,
    sector_df:   pd.DataFrame,
    trades:      pd.DataFrame,
    source_run:  str,
    out_path:    Path,
) -> None:
    """
    Write a 5-sheet investor-grade Excel workbook.

    Sheet 1 — Executive Summary    : all 7 metric groups, two-column layout
    Sheet 2 — Monthly Returns      : year × month heat-map table
    Sheet 3 — Sector Breakdown     : per-sector metrics
    Sheet 4 — Daily Equity Curve   : raw equity, drawdown, open positions
    Sheet 5 — Trade Log            : full trade-level record

    Colour scheme consistent with 09D/09E outputs.
    """
    try:
        from openpyxl.styles import (
            PatternFill, Font, Alignment, Border, Side,
        )
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("  [WARNING] openpyxl not installed — Excel skipped.")
        return

    # ── Shared style objects ──────────────────────────────────────────────────
    thin   = Side(style="thin", color=MID_GREY)
    bdr    = Border(left=thin, right=thin, top=thin, bottom=thin)

    def fill(hex_col: str) -> PatternFill:
        return PatternFill("solid", fgColor=hex_col)

    def font(bold=False, size=10, color=DARK_GREY, italic=False) -> Font:
        return Font(bold=bold, size=size, color=color,
                    italic=italic, name="Arial")

    HDR_FILL  = fill(NAVY)
    HDR_FONT  = font(bold=True, color="FFFFFF")
    SEC_FILL  = fill(BLUE)
    SEC_FONT  = font(bold=True, color="FFFFFF")
    TITLE_F   = font(bold=True, size=14, color=NAVY)
    SUB_F     = font(size=10,  italic=True, color="606060")
    META_F    = font(size=9,   italic=True, color="909090")
    BOLD_F    = font(bold=True)
    NORM_F    = font()
    GREEN_F   = font(bold=True, color=D_GREEN)
    RED_F     = font(bold=True, color=D_RED)
    LEFT_A    = Alignment(horizontal="left",   vertical="center", wrap_text=False)
    RIGHT_A   = Alignment(horizontal="right",  vertical="center")
    CENTER_A  = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def _apply_hdr(ws, row: int, labels: list, widths: list):
        ws.row_dimensions[row].height = 26
        for c, (lab, w) in enumerate(zip(labels, widths), 1):
            cell = ws.cell(row=row, column=c, value=lab)
            cell.fill      = HDR_FILL
            cell.font      = HDR_FONT
            cell.alignment = CENTER_A
            cell.border    = bdr
            ws.column_dimensions[get_column_letter(c)].width = w

    def _data_row(ws, row: int, values: list, fills: list, bold=False, text_cols=()):
        ws.row_dimensions[row].height = 17
        for c, (v, fhex) in enumerate(zip(values, fills), 1):
            raw  = v.item() if hasattr(v, "item") else v
            cell = ws.cell(row=row, column=c, value=raw)
            cell.fill      = fill(fhex)
            cell.font      = font(bold=bold)
            cell.border    = bdr
            cell.alignment = LEFT_A if c in text_cols else RIGHT_A

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:

        # ══════════════════════════════════════════════════════════════════════
        # SHEET 1 — EXECUTIVE SUMMARY
        # Two-column layout: each group on left (label/value) and right (label/value)
        # ══════════════════════════════════════════════════════════════════════
        writer.book.create_sheet("Executive Summary")
        ws1 = writer.book["Executive Summary"]

        # -- Title block (rows 1–3) -------------------------------------------
        ws1.merge_cells("A1:F1")
        ws1["A1"] = "ALGO-STOCKS — Full Universe Portfolio Baseline"
        ws1["A1"].font = TITLE_F;  ws1["A1"].alignment = LEFT_A
        ws1.row_dimensions[1].height = 30

        ws1.merge_cells("A2:F2")
        ws1["A2"] = (f"Methodology: virtual parallel accounts  |  "
                     f"${VIRTUAL_ACCOUNT_PER_TRADE:,.0f} per trade  |  "
                     f"Zero blocking  |  "
                     f"{metrics['total_trades']:,} trades  |  "
                     f"{metrics['tickers_included']:,} tickers")
        ws1["A2"].font = SUB_F;  ws1["A2"].alignment = LEFT_A

        ws1.merge_cells("A3:F3")
        ws1["A3"] = (f"Source run: {source_run}    |    "
                     f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
        ws1["A3"].font = META_F;  ws1["A3"].alignment = LEFT_A
        ws1.row_dimensions[3].height = 14

        # Column widths: A(label), B(value), C(spacer), D(label), E(value), F(note)
        for col, w in zip(["A","B","C","D","E","F"],
                           [32,  20,  2,  32,  20,  42]):
            ws1.column_dimensions[col].width = w

        # -- Helper: write one label/value pair with directional colouring -----
        def _put(row: int, col_l: int, col_v: int,
                 label: str, value,
                 row_fill: str = L_GREY,
                 direction: str = None,   # "pos_good" | "neg_bad" | None
                 note: str = ""):
            ws1.row_dimensions[row].height = 17
            cl = ws1.cell(row=row, column=col_l, value=label)
            cv = ws1.cell(row=row, column=col_v, value=value)
            cl.fill = cv.fill = fill(row_fill)
            cl.font  = BOLD_F;  cv.font = NORM_F
            cl.border = cv.border = bdr
            cl.alignment = LEFT_A;  cv.alignment = RIGHT_A
            # Directional colouring on the value cell
            if direction and isinstance(value, (int, float)):
                is_pos = value > 0
                if direction == "pos_good":
                    cv.font = GREEN_F if is_pos else RED_F
                elif direction == "neg_bad":
                    cv.font = RED_F   if not is_pos else NORM_F
            if note:
                cn = ws1.cell(row=row, column=6, value=note)
                cn.font = font(size=9, italic=True, color="909090")
                cn.alignment = LEFT_A

        # -- Helper: section header spanning 5 columns -------------------------
        def _sec(row: int, label_left: str, label_right: str = ""):
            for col, lab in [(1, label_left), (4, label_right)]:
                for cc in range(col, col + 2):
                    cell = ws1.cell(row=row, column=cc,
                                    value=lab if cc == col else "")
                    cell.fill      = fill(NAVY)
                    cell.font      = font(bold=True, color="FFFFFF")
                    cell.border    = bdr
                    cell.alignment = LEFT_A
            ws1.row_dimensions[row].height = 20

        m   = metrics
        row = 5   # start below title block

        # ── A. Returns  /  B. Risk ────────────────────────────────────────────
        _sec(row, "A.  RETURN METRICS", "B.  RISK METRICS"); row += 1

        ret_risk_pairs = [
            # left label, left value, left_dir, right label, right value, right_dir, fill, note_left
            ("Net PnL ($)",
             f"${m['net_pnl_usd']:,.0f}",
             None,
             "Max Drawdown %",
             m["max_drawdown_pct"],
             "neg_bad",
             L_GREY,
             "Primary investor concern"),
            ("Ann. Return on Avg Deployed %",
             m["ann_ret_on_avg_deployed_pct"],
             "pos_good",
             "Max DD Duration (days)",
             m["max_dd_duration_days"],
             None,
             WHITE,
             "Primary return metric"),
            ("Return on Avg Deployed % (total)",
             m["ret_on_avg_deployed_pct"],
             "pos_good",
             "Recovery Time (days)",
             m["recovery_days"] if m["recovery_days"] >= 0 else "Not yet recovered",
             None,
             L_GREY,
             ""),
            ("CAGR % (total virtual capital base)",
             m["cagr_total_base_pct"],
             "pos_good",
             "Avg Drawdown %",
             m["avg_drawdown_pct"],
             "neg_bad",
             WHITE,
             "Conservative full-base view"),
            ("Net Return % (total virtual base)",
             m["net_return_total_base_pct"],
             "pos_good",
             "Ann. Volatility %",
             m["ann_volatility_pct"],
             None,
             L_GREY,
             "Informational — see note below"),
            ("Gross Wins ($)",
             f"${m['gross_wins_usd']:,.0f}",
             None,
             "VaR 95% (daily)",
             f"{m['var_95_pct']:.2f}%",
             None,
             WHITE,
             ""),
            ("Gross Losses ($)",
             f"${m['gross_losses_usd']:,.0f}",
             None,
             "CVaR 95% / Exp. Shortfall",
             f"{m['cvar_95_pct']:.2f}%",
             None,
             L_GREY,
             ""),
            ("",
             "",
             None,
             "Pain Index",
             m["pain_index"],
             None,
             WHITE,
             "Avg drawdown magnitude all days"),
        ]
        for ll, lv, ld, rl, rv, rd, bg, note in ret_risk_pairs:
            _put(row, 1, 2, ll, lv, bg, ld, note)
            _put(row, 4, 5, rl, rv, bg, rd)
            row += 1

        row += 1
        # ── C. Risk-adjusted  /  D. Trade Statistics ─────────────────────────
        _sec(row, "C.  RISK-ADJUSTED RATIOS", "D.  TRADE STATISTICS"); row += 1

        ra_trade_pairs = [
            ("Sharpe Ratio (annualised)",
             m["sharpe_ratio"],        "pos_good",
             "Total Trades",           f"{m['total_trades']:,}",    None, L_GREY, ""),
            ("Sortino Ratio (annualised)",
             m["sortino_ratio"],       "pos_good",
             "Win Rate %",             m["win_rate_pct"],           None, WHITE,  ""),
            ("Calmar Ratio",
             m["calmar_ratio"],        "pos_good",
             "Profit Factor",          m["profit_factor"],          "pos_good", L_GREY, ""),
            ("Omega Ratio",
             m["omega_ratio"],         "pos_good",
             "Expectancy R",           m["expectancy_r"],           "pos_good", WHITE,  ""),
            ("",  "", None,
             "Payoff Ratio (avg W / avg L R)", m["payoff_ratio"],   "pos_good", L_GREY, ""),
            ("",  "", None,
             "Avg Win %",              m["avg_win_pct"],            "pos_good", WHITE,  ""),
            ("",  "", None,
             "Avg Loss %",             m["avg_loss_pct"],           "neg_bad",  L_GREY, ""),
            ("",  "", None,
             "Best Trade %",           m["best_trade_pct"],         "pos_good", WHITE,  ""),
            ("",  "", None,
             "Worst Trade %",          m["worst_trade_pct"],        "neg_bad",  L_GREY, ""),
            ("",  "", None,
             "Avg Hold Days",          m["avg_hold_days"],          None,       WHITE,  ""),
            ("",  "", None,
             "Max Hold Days",          m["max_hold_days"],          None,       L_GREY, ""),
            ("",  "", None,
             "Max Consec. Wins",       m["max_consec_wins"],        None,       WHITE,  ""),
            ("",  "", None,
             "Max Consec. Losses",     m["max_consec_losses"],      None,       L_GREY, ""),
        ]
        for ll, lv, ld, rl, rv, rd, bg, note in ra_trade_pairs:
            _put(row, 1, 2, ll, lv, bg, ld, note)
            _put(row, 4, 5, rl, rv, bg, rd)
            row += 1

        row += 1
        # ── E. Consistency  /  F. Position & Deployment ───────────────────────
        _sec(row, "E.  CONSISTENCY", "F.  POSITION & DEPLOYMENT"); row += 1

        con_pos_pairs = [
            ("% Positive Months",         m["pct_positive_months"],    None,
             "Avg Open Positions",         m["avg_open_positions"],      None, L_GREY),
            ("% Positive Quarters",        m["pct_positive_quarters"],  None,
             "Max Open Positions",         m["max_open_positions"],      None, WHITE),
            ("% Positive Years",           m["pct_positive_years"],     None,
             "Avg Deployed Capital ($)",   f"${m['avg_deployed_capital']:,.0f}",  None, L_GREY),
            ("Best Month %",               m["best_month_pct"],         "pos_good",
             "Peak Deployed Capital ($)",  f"${m['peak_deployed_capital']:,.0f}", None, WHITE),
            ("Worst Month %",              m["worst_month_pct"],        "neg_bad",
             "Avg Utilisation %",          m["avg_utilisation_pct"],    None, L_GREY),
            ("Avg Month %",                m["avg_month_pct"],          "pos_good",
             "Peak Utilisation %",         m["peak_utilisation_pct"],   None, WHITE),
        ]
        for ll, lv, ld, rl, rv, rd, bg in con_pos_pairs:
            _put(row, 1, 2, ll, lv, bg, ld)
            _put(row, 4, 5, rl, rv, bg, rd)
            row += 1

        row += 1
        # ── G. Exit Quality ───────────────────────────────────────────────────
        _sec(row, "G.  EXIT QUALITY  (how positions are closed)", ""); row += 1

        for reason, data in sorted(
            m["exit_reasons"].items(), key=lambda x: -x[1]["count"]
        ):
            label   = reason.replace("_", " ").title()
            val_str = f"{data['count']:,}  ({data['pct']:.1f}%)"
            if "stage9" in reason:
                bg = L_GREEN
            elif "stop" in reason:
                bg = L_RED
            elif "time" in reason:
                bg = AMBER
            else:
                bg = L_GREY
            _put(row, 1, 2, label, val_str, bg, None)
            row += 1

        row += 2
        # ── Methodology note ──────────────────────────────────────────────────
        ws1.merge_cells(f"A{row}:F{row}")
        note_cell = ws1.cell(row=row, column=1,
            value=("NOTE — Return denominator:  "
                   "Ann. Return on Avg Deployed uses capital actually working at each moment "
                   "(avg_open_positions × $10k) — this is the correct comparison to 09C.  "
                   "Net Return % and CAGR % use total virtual capital ($10k × n_trades) "
                   "as the base — not a realistic deployment figure, shown for completeness only."))
        note_cell.font      = font(size=9, italic=True, color="808080")
        note_cell.alignment = Alignment(wrap_text=True)
        ws1.row_dimensions[row].height = 36

        ws1.freeze_panes = ws1.cell(row=5, column=1)

        # ══════════════════════════════════════════════════════════════════════
        # SHEET 2 — MONTHLY RETURNS (heat-map table)
        # ══════════════════════════════════════════════════════════════════════
        sheet2_name = "Monthly Returns"
        monthly_out = monthly_tbl.reset_index()
        monthly_out.to_excel(writer, sheet_name=sheet2_name, index=False)
        ws2 = writer.sheets[sheet2_name]

        m_cols   = monthly_out.columns.tolist()
        m_widths = [8] + [9] * (len(m_cols) - 1)
        _apply_hdr(ws2, 1, m_cols, m_widths)

        # Colour each return cell: green positive, red negative, grey blank
        for r_idx in range(2, ws2.max_row + 1):
            # Year column
            yc = ws2.cell(row=r_idx, column=1)
            yc.fill = fill(L_GREY); yc.font = BOLD_F
            yc.border = bdr; yc.alignment = CENTER_A
            # Month / annual columns
            for c_idx in range(2, len(m_cols) + 1):
                cell = ws2.cell(row=r_idx, column=c_idx)
                v    = cell.value
                cell.border    = bdr
                cell.alignment = RIGHT_A
                if isinstance(v, (int, float)) and v == v:  # not NaN
                    if v > 0:
                        cell.fill = fill(PALE_GREEN)
                        cell.font = Font(bold=True, color=D_GREEN, name="Arial", size=10)
                    elif v < 0:
                        cell.fill = fill(PALE_RED)
                        cell.font = Font(bold=True, color=D_RED,   name="Arial", size=10)
                    else:
                        cell.fill = fill(L_GREY);  cell.font = NORM_F
                else:
                    cell.fill  = fill(L_GREY)
                    cell.value = "—"
                    cell.font  = font(color="AAAAAA")
        ws2.freeze_panes = ws2.cell(row=2, column=2)

        # ══════════════════════════════════════════════════════════════════════
        # SHEET 3 — SECTOR BREAKDOWN
        # ══════════════════════════════════════════════════════════════════════
        if not sector_df.empty:
            sector_df.to_excel(writer, sheet_name="Sector Breakdown", index=False)
            ws3  = writer.sheets["Sector Breakdown"]
            s_cols   = sector_df.columns.tolist()
            s_widths = [26, 10, 13, 15, 14, 16, 12, 12, 15]
            _apply_hdr(ws3, 1, s_cols, s_widths[:len(s_cols)])
            for r_idx in range(2, ws3.max_row + 1):
                bg = L_GREY if r_idx % 2 == 0 else WHITE
                # Colour net PnL column (col 6) green/red
                pnl_col = s_cols.index("Net PnL ($)") + 1
                pnl_val = ws3.cell(row=r_idx, column=pnl_col).value
                for c_idx in range(1, len(s_cols) + 1):
                    cell       = ws3.cell(row=r_idx, column=c_idx)
                    cell.fill  = fill(bg)
                    cell.font  = NORM_F
                    cell.border = bdr
                    cell.alignment = LEFT_A if c_idx == 1 else RIGHT_A
                # Override PnL cell colour
                pnl_cell = ws3.cell(row=r_idx, column=pnl_col)
                if isinstance(pnl_val, (int, float)):
                    pnl_cell.font = GREEN_F if pnl_val > 0 else RED_F
            ws3.freeze_panes = ws3.cell(row=2, column=1)

        # ══════════════════════════════════════════════════════════════════════
        # SHEET 4 — DAILY EQUITY CURVE
        # ══════════════════════════════════════════════════════════════════════
        eq_export = daily_eq.copy()
        eq_export["date"] = eq_export["date"].dt.strftime("%Y-%m-%d")
        eq_export["daily_return_pct"] = (
            pd.Series(daily_eq["equity"].values).pct_change() * 100
        ).round(4).values
        eq_export["drawdown_pct"] = (
            (daily_eq["equity"]
             - daily_eq["equity"].cummax())
            / daily_eq["equity"].cummax() * 100
        ).round(4).values

        cols4   = ["date","equity","open_positions","deployed_capital",
                   "cumulative_pnl","daily_return_pct","drawdown_pct"]
        hdrs4   = ["Date","Portfolio Equity ($)","Open Positions",
                   "Deployed Capital ($)","Cumulative PnL ($)",
                   "Daily Return %","Drawdown %"]
        wids4   = [14, 22, 16, 22, 22, 16, 14]

        eq_export[cols4].to_excel(writer, sheet_name="Daily Equity Curve", index=False)
        ws4 = writer.sheets["Daily Equity Curve"]
        _apply_hdr(ws4, 1, hdrs4, wids4)
        for r_idx in range(2, ws4.max_row + 1):
            bg = L_GREY if r_idx % 2 == 0 else WHITE
            for c_idx in range(1, len(cols4) + 1):
                cell       = ws4.cell(row=r_idx, column=c_idx)
                cell.fill  = fill(bg)
                cell.font  = NORM_F
                cell.border = bdr
                cell.alignment = LEFT_A if c_idx == 1 else RIGHT_A
        ws4.freeze_panes = ws4.cell(row=2, column=1)

        # ══════════════════════════════════════════════════════════════════════
        # SHEET 5 — TRADE LOG
        # ══════════════════════════════════════════════════════════════════════
        t_log = trades[[
            "ticker","sector","signal_type",
            "entry_date","exit_date","entry_price","exit_price",
            "exit_reason","hold_days","pnl_dollar","pnl_pct","pnl_r",
        ]].copy()
        t_log["entry_date"] = t_log["entry_date"].dt.strftime("%Y-%m-%d")
        t_log["exit_date"]  = t_log["exit_date"].dt.strftime("%Y-%m-%d")
        t_log["pnl_pct"]    = t_log["pnl_pct"].round(4)

        t_log.to_excel(writer, sheet_name="Trade Log", index=False)
        ws5 = writer.sheets["Trade Log"]
        t_hdrs  = ["Ticker","Sector","Signal","Entry Date","Exit Date",
                   "Entry $","Exit $","Exit Reason","Hold Days",
                   "PnL ($)","PnL %","PnL R"]
        t_wids  = [10, 24, 14, 13, 13, 12, 12, 18, 12, 13, 10, 10]
        _apply_hdr(ws5, 1, t_hdrs, t_wids)
        for r_idx in range(2, ws5.max_row + 1):
            pnl_val = ws5.cell(row=r_idx, column=10).value
            if isinstance(pnl_val, (int, float)):
                row_fill = L_GREEN if float(pnl_val) > 0 else L_RED
            else:
                row_fill = WHITE
            for c_idx in range(1, len(t_hdrs) + 1):
                cell       = ws5.cell(row=r_idx, column=c_idx)
                cell.fill  = fill(row_fill)
                cell.font  = NORM_F
                cell.border = bdr
                cell.alignment = LEFT_A if c_idx <= 3 else RIGHT_A
        ws5.freeze_panes = ws5.cell(row=2, column=1)

        # ── Move Executive Summary to tab 1 ───────────────────────────────────
        wb = writer.book
        wb.move_sheet("Executive Summary", offset=-(len(wb.sheetnames) - 1))

    print(f"  Excel written         : {out_path}")


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

def main() -> None:
    print("=" * 70)
    print("09F — FULL UNIVERSE VIRTUAL PORTFOLIO (ZERO-BLOCKING BASELINE)")
    print("=" * 70)
    ts_start = datetime.now()

    # Load config ──────────────────────────────────────────────────────────────
    cfg            = _load_cfg(BACKTEST_CFG)
    source_run_tag = str(cfg.get("batches", {}).get("source_run_tag", ""))
    if not source_run_tag:
        print("[ERROR] batches.source_run_tag not set in config/backtest.yaml")
        sys.exit(1)

    print(f"Source run tag        : {source_run_tag}")
    print(f"Virtual acct/trade    : ${VIRTUAL_ACCOUNT_PER_TRADE:,.0f}")
    print(f"Use filtered tickers  : {USE_FILTERED_TICKERS}")

    # Load trades ──────────────────────────────────────────────────────────────
    trades_path = BACKTESTS_DIR / source_run_tag / "universe" / "trades_all.parquet"
    if not trades_path.exists():
        print(f"\n[ERROR] Not found: {trades_path}")
        sys.exit(1)

    trades = pd.read_parquet(trades_path)
    trades["entry_date"] = pd.to_datetime(trades["entry_date"])
    trades["exit_date"]  = pd.to_datetime(trades["exit_date"])
    print(f"\nLoaded                : {len(trades):,} trades from "
          f"{trades['ticker'].nunique():,} tickers")

    # Sector enrichment ────────────────────────────────────────────────────────
    if "spider_id" not in trades.columns:
        if MEMBERSHIPS.exists():
            mem = pd.read_csv(MEMBERSHIPS)[["ticker", "spider_id"]].drop_duplicates("ticker")
            trades = trades.merge(mem, on="ticker", how="left")
        else:
            trades["spider_id"] = None
    trades["sector"] = trades["spider_id"].apply(_sector_from_spider_id)

    # Optional ticker filter ───────────────────────────────────────────────────
    if USE_FILTERED_TICKERS:
        ft_path = FILTER_DIR / "filtered_tickers.csv"
        if not ft_path.exists():
            print("[ERROR] filtered_tickers.csv not found — run 09D first "
                  "or set USE_FILTERED_TICKERS=False")
            sys.exit(1)
        passing = set(pd.read_csv(ft_path)["ticker"].tolist())
        before  = len(trades)
        trades  = trades[trades["ticker"].isin(passing)].copy()
        print(f"Ticker filter         : {len(trades):,} kept "
              f"({before - len(trades):,} removed)")

    # Confirm scope ────────────────────────────────────────────────────────────
    total_virtual_capital = len(trades) * VIRTUAL_ACCOUNT_PER_TRADE
    print(f"\nTrades in simulation  : {len(trades):,}  (ALL execute — zero blocking)")
    print(f"Unique tickers        : {trades['ticker'].nunique():,}")
    print(f"Date range            : "
          f"{trades['entry_date'].min().date()} → {trades['exit_date'].max().date()}")
    print(f"Total virtual capital : ${total_virtual_capital:,.0f}  "
          f"({len(trades):,} × ${VIRTUAL_ACCOUNT_PER_TRADE:,.0f})")

    # Build equity curve ───────────────────────────────────────────────────────
    # reference_capital is set inside build_daily_equity_curve based on
    # avg_open_positions × $10k — the capital actually deployed on a typical day
    daily_eq = build_daily_equity_curve(trades)
    start_equity = float(daily_eq["reference_capital"].iloc[0])
    print(f"Reference capital     : ${start_equity:,.0f}  "
          f"(avg deployed = {start_equity / VIRTUAL_ACCOUNT_PER_TRADE:.0f} avg positions × $10k)")

    # Compute full metric suite ────────────────────────────────────────────────
    print("\nComputing metric suite ...")
    metrics = compute_all_metrics(daily_eq, trades, start_equity)

    # Build supporting tables ──────────────────────────────────────────────────
    monthly_tbl = build_monthly_table(daily_eq)
    sector_df   = build_sector_breakdown(trades)

    # Write outputs ────────────────────────────────────────────────────────────
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    daily_eq.to_parquet(REPORTS_DIR / "daily_equity.parquet", index=False)
    monthly_tbl.reset_index().to_csv(REPORTS_DIR / "monthly_returns.csv")

    tc = trades[[
        "ticker","sector","entry_date","exit_date","exit_reason",
        "hold_days","pnl_dollar","pnl_pct","pnl_r","signal_type",
    ]].copy()
    tc["virtual_account"] = VIRTUAL_ACCOUNT_PER_TRADE
    tc["final_value"]     = VIRTUAL_ACCOUNT_PER_TRADE + tc["pnl_dollar"]
    tc.to_csv(REPORTS_DIR / "trade_contributions.csv", index=False)

    elapsed = (datetime.now() - ts_start).total_seconds()
    report  = {
        "generated_at":    datetime.now().isoformat(),
        "source_run_tag":  source_run_tag,
        "methodology":     "virtual_parallel_accounts",
        "virtual_acct":    VIRTUAL_ACCOUNT_PER_TRADE,
        "use_filtered":    USE_FILTERED_TICKERS,
        "elapsed_seconds": round(elapsed, 1),
        "metrics":         metrics,
        "config_snapshot": cfg,
    }
    with (REPORTS_DIR / "portfolio_summary.json").open("w", encoding="utf-8") as f:
        json.dump(report, f, indent=2, default=str)

    write_investor_excel(
        metrics, daily_eq, monthly_tbl, sector_df, trades,
        source_run_tag, REPORTS_DIR / "full_universe_report.xlsx",
    )

    # Console summary ──────────────────────────────────────────────────────────
    m = metrics
    W = 65
    print(f"\n{'═'*W}")
    print(f"  FULL UNIVERSE PORTFOLIO — ZERO BLOCKING BASELINE")
    print(f"{'═'*W}")
    print(f"  Trades executed (100%)             : {m['total_trades']:>10,}")
    print(f"  Unique tickers                     : {m['tickers_included']:>10,}")
    print(f"{'─'*W}")
    print(f"  A. RETURNS")
    print(f"  Net PnL ($)                        : ${m['net_pnl_usd']:>14,.0f}")
    print(f"  Ann. Return on Avg Deployed %      : {m['ann_ret_on_avg_deployed_pct']:>10.2f}%  ← primary")
    print(f"  Return on Avg Deployed % (total)   : {m['ret_on_avg_deployed_pct']:>10.2f}%")
    print(f"  CAGR % (total virtual base)        : {m['cagr_total_base_pct']:>10.2f}%")
    print(f"  Net Return % (total virtual base)  : {m['net_return_total_base_pct']:>10.2f}%")
    print(f"{'─'*W}")
    print(f"  B. RISK")
    print(f"  Max Drawdown %                     : {m['max_drawdown_pct']:>10.2f}%")
    print(f"  Max DD Duration (days)             : {m['max_dd_duration_days']:>10,}")
    print(f"  Recovery Time (days)               : "
          f"{'Not yet recovered' if m['recovery_days'] < 0 else str(m['recovery_days']):>10}")
    print(f"  Ann. Volatility %                  : {m['ann_volatility_pct']:>10.2f}%")
    print(f"  VaR 95% (daily)                    : {m['var_95_pct']:>10.2f}%")
    print(f"  CVaR 95% / Exp. Shortfall          : {m['cvar_95_pct']:>10.2f}%")
    print(f"{'─'*W}")
    print(f"  C. RISK-ADJUSTED")
    print(f"  Sharpe Ratio (ann.)                : {m['sharpe_ratio']:>10.4f}")
    print(f"  Sortino Ratio (ann.)               : {m['sortino_ratio']:>10.4f}")
    print(f"  Calmar Ratio                       : {m['calmar_ratio']:>10.4f}")
    print(f"  Omega Ratio                        : {m['omega_ratio']:>10.4f}")
    print(f"{'─'*W}")
    print(f"  D. TRADE STATISTICS")
    print(f"  Win Rate %                         : {m['win_rate_pct']:>10.1f}%")
    print(f"  Profit Factor                      : {str(m['profit_factor']):>10}")
    print(f"  Expectancy R                       : {m['expectancy_r']:>10.4f}")
    print(f"  Payoff Ratio (avg W / avg L)       : {m['payoff_ratio']:>10.4f}")
    print(f"  Avg Win %  /  Avg Loss %           : "
          f"{m['avg_win_pct']:>6.2f}%  /  {m['avg_loss_pct']:.2f}%")
    print(f"  Max Consecutive Wins / Losses      : "
          f"{m['max_consec_wins']:>5} / {m['max_consec_losses']}")
    print(f"{'─'*W}")
    print(f"  E. CONSISTENCY")
    print(f"  % Positive Months                  : {m['pct_positive_months']:>10.1f}%")
    print(f"  % Positive Quarters                : {m['pct_positive_quarters']:>10.1f}%")
    print(f"  % Positive Years                   : {m['pct_positive_years']:>10.1f}%")
    print(f"  Best / Worst Month                 : "
          f"{m['best_month_pct']:>5.2f}%  /  {m['worst_month_pct']:.2f}%")
    print(f"{'─'*W}")
    print(f"  F. POSITIONS")
    print(f"  Avg / Max Open Positions           : "
          f"{m['avg_open_positions']:>5.1f}  /  {m['max_open_positions']}")
    print(f"  Avg Deployed Capital ($)           : ${m['avg_deployed_capital']:>13,.0f}")
    print(f"  Avg Utilisation %                  : {m['avg_utilisation_pct']:>10.1f}%")
    print(f"{'─'*W}")
    print(f"  G. EXIT QUALITY")
    for reason, data in sorted(
        m["exit_reasons"].items(), key=lambda x: -x[1]["count"]
    ):
        print(f"  {reason:<38}: "
              f"{data['count']:>7,}  ({data['pct']:.1f}%)")
    print(f"{'═'*W}")
    print(f"\nOutputs → {REPORTS_DIR}")
    print(f"  full_universe_report.xlsx  ← open this")
    print(f"  daily_equity.parquet  |  monthly_returns.csv")
    print(f"  trade_contributions.csv  |  portfolio_summary.json")
    print(f"\nElapsed               : {elapsed:.1f}s")
    print("=" * 70)
    print("09F COMPLETE")
    print("=" * 70)


if __name__ == "__main__":
    main()
